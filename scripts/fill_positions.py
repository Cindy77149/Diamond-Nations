"""
Fill missing POS (X / None) from Wikipedia player infobox.
- Catchers  → C  (direct from section)
- DH        → DH (direct from section)
- Infielders / Outfielders → look up Wikipedia position field
"""
from __future__ import annotations
import re, json, time, urllib.request, urllib.parse
from pathlib import Path
from openpyxl import load_workbook

OUTPUT = Path(__file__).parent.parent / "wbc_roster.xlsx"
HEADERS = {"User-Agent": "WBC-Stats-Bot/1.0 (baseball stats research)"}

# ── Wikipedia position text → standard abbreviation ─────────────────────────
POS_MAP = {
    "pitcher":          "P",
    "starting pitcher": "SP",
    "relief pitcher":   "RP",
    "closer":           "CP",
    "catcher":          "C",
    "first base":       "1B",  "first baseman":  "1B",
    "second base":      "2B",  "second baseman": "2B",
    "third base":       "3B",  "third baseman":  "3B",
    "shortstop":        "SS",
    "left field":       "LF",  "left fielder":   "LF",
    "center field":     "CF",  "center fielder": "CF",
    "right field":      "RF",  "right fielder":  "RF",
    "outfield":         "OF",  "outfielder":     "OF",
    "infield":          "INF", "infielder":      "INF",
    "designated hitter":"DH",
    "utility":          "UT",
}

# Section → fallback when Wikipedia also fails
SECTION_FALLBACK = {
    "Catchers":          "C",
    "Infielders":        "INF",
    "Outfielders":       "OF",
    "Designated Hitter": "DH",
    "Pitchers":          "P",
    "Two-Way Players":   "SP/RP",
}


def _api_get(url: str) -> dict:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=12) as r:
        return json.loads(r.read())


def _name_matches(en_name: str, page_title: str) -> bool:
    ep = set(en_name.lower().replace("-", " ").split())
    tp = set(page_title.lower().replace("-", " ").split())
    return len(ep & tp) >= min(2, len(ep))


def parse_position(wikitext: str) -> str:
    """Extract position from infobox wikitext."""
    # Try | position = ... or | pos = ...
    for pat in [r"\|\s*position\s*=\s*([^\n|]+)", r"\|\s*pos\s*=\s*([^\n|]+)"]:
        m = re.search(pat, wikitext, re.IGNORECASE)
        if m:
            raw = re.sub(r"\[\[([^\]|]+\|)?([^\]]+)\]\]", r"\2", m.group(1))
            raw = re.sub(r"\{\{[^}]*\}\}", "", raw).strip().lower()
            for key, abbr in POS_MAP.items():
                if key in raw:
                    return abbr
    # Fallback: scan full infobox area (first 2000 chars) for position keywords
    lower_wt = wikitext[:2000].lower()
    for key, abbr in POS_MAP.items():
        if re.search(r"\b" + re.escape(key) + r"\b", lower_wt):
            return abbr
    return ""


def lookup_position(en_name: str) -> str:
    search_url = (
        "https://en.wikipedia.org/w/api.php?action=query&list=search"
        f"&srsearch={urllib.parse.quote(en_name + ' baseball')}"
        "&srlimit=3&format=json"
    )
    try:
        data = _api_get(search_url)
        results = data.get("query", {}).get("search", [])
        title = next((r["title"] for r in results if _name_matches(en_name, r["title"])), "")
        if not title:
            return ""
        parse_url = (
            "https://en.wikipedia.org/w/api.php?action=parse"
            f"&page={urllib.parse.quote(title)}&prop=wikitext&section=0&format=json"
        )
        data2 = _api_get(parse_url)
        wikitext = data2.get("parse", {}).get("wikitext", {}).get("*", "")
        # Verify it's a baseball player
        if not any(kw in wikitext[:2000].lower() for kw in
                   ("baseball", "pitcher", "baseman", "shortstop", "outfield",
                    "catcher", "npb", "kbo", "mlb", "cpbl")):
            return ""
        return parse_position(wikitext)
    except Exception:
        return ""


def main() -> None:
    wb = load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    # Collect rows needing position lookup
    needs = []
    for r in range(2, ws.max_row + 1):
        pos = ws.cell(r, col["POS"]).value
        if pos not in (None, "X"):
            continue
        sec  = ws.cell(r, col["section"]).value
        name = ws.cell(r, col["player_name"]).value
        if name:
            needs.append((r, name, sec))

    print(f"Rows needing position: {len(needs)}")

    # Deduplicate by player name
    unique: dict[str, str] = {}  # name → pos
    for _, name, sec in needs:
        if name not in unique:
            # Direct from section for Catchers / DH
            if sec in ("Catchers", "Designated Hitter"):
                unique[name] = SECTION_FALLBACK[sec]
            else:
                unique[name] = ""  # needs lookup

    to_lookup = [(name, sec) for name, sec in
                 [(n, next(s for _, n2, s in needs if n2 == n)) for n in unique if not unique[n]]
                 ]
    # Deduplicate
    seen = set()
    to_lookup_dedup = []
    for name, sec in to_lookup:
        if name not in seen:
            seen.add(name)
            to_lookup_dedup.append((name, sec))

    print(f"Direct from section: {sum(1 for v in unique.values() if v)}")
    print(f"Wikipedia lookups: {len(to_lookup_dedup)}")

    for i, (name, sec) in enumerate(to_lookup_dedup):
        pos = lookup_position(name)
        if not pos:
            pos = SECTION_FALLBACK.get(sec, "X")
        unique[name] = pos
        print(f"  [{i+1}/{len(to_lookup_dedup)}] {name} ({sec}) → {pos}")
        time.sleep(0.35)

    # Write back
    filled = 0
    for r, name, sec in needs:
        new_pos = unique.get(name, SECTION_FALLBACK.get(sec, "X"))
        ws.cell(r, col["POS"]).value = new_pos
        filled += 1

    wb.save(OUTPUT)
    print(f"\nFilled {filled} positions.")


if __name__ == "__main__":
    main()
