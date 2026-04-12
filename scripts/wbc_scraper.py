"""
WBC scraper — unified script.

Usage:
  python wbc_scraper.py stats   # scrape batting/pitching stats for all teams
  python wbc_scraper.py roster  # fill jersey/bats/throws/height/weight from roster pages
  python wbc_scraper.py all     # run both (stats first, then roster)

Optional filters:
  --teams  japan korea ...       only process these nationality keys
  --years  2026 2023 ...         only process these years
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, Page

# ── Config ────────────────────────────────────────────────────────────────────

OUTPUT = Path(__file__).parent.parent / "wbc_roster.xlsx"
NO_DATA = "No roster data is available"

# (slug, nationality, team_filter_name, stats_years, roster_years)
TEAMS: list[tuple[str, str, str, list[int], list[int]]] = [
    ("chinese-taipei", "台灣",    "Chinese Taipei", [2026, 2023, 2017, 2013, 2006],       []),
    ("japan",          "日本",    "Japan",           [2026, 2023, 2017, 2013, 2009, 2006], [2026, 2023, 2017, 2013, 2009, 2006]),
    ("korea",          "韓國",    "Korea",            [2026, 2023, 2017, 2013, 2009, 2006], [2026, 2023, 2017, 2013, 2009, 2006]),
    ("united-states",  "美國",    "United States",    [2026, 2023, 2017, 2013, 2009, 2006], [2026, 2023, 2017, 2013, 2009, 2006]),
    ("dominican-republic", "多明尼加", "Dominican Republic", [2026, 2023, 2017, 2013, 2009, 2006], [2026, 2023, 2017, 2013, 2009, 2006]),
    ("venezuela",      "委內瑞拉", "Venezuela",        [2026, 2023, 2017, 2013, 2009, 2006], [2026, 2023, 2017, 2013, 2009, 2006]),
    ("cuba",           "古巴",    "Cuba",             [2026, 2023, 2017, 2013, 2009, 2006], [2026, 2023, 2017, 2013, 2009, 2006]),
    ("puerto-rico",    "波多黎各", "Puerto Rico",      [2026, 2023, 2017, 2013, 2009, 2006], [2026, 2023, 2017, 2013, 2009, 2006]),
]

POS_TO_SECTION = {
    "P": "Pitchers", "SP": "Pitchers", "RP": "Pitchers",
    "C": "Catchers",
    "1B": "Infielders", "2B": "Infielders", "3B": "Infielders",
    "SS": "Infielders", "IF": "Infielders",
    "LF": "Outfielders", "CF": "Outfielders", "RF": "Outfielders", "OF": "Outfielders",
    "DH": "Designated Hitter",
}

STATS_BASE_URL  = "https://www.mlb.com/world-baseball-classic/stats/{slug}"
ROSTER_BASE_URL = "https://www.mlb.com/world-baseball-classic/roster/{slug}?season={year}"

# ── Shared JS extractors ───────────────────────────────────────────────────────

STATS_JS = """
() => {
    const clean = s => (s || '').replace(/\\s+/g, ' ').trim();
    const table = document.querySelector('table');
    if (!table) return null;
    const headers = Array.from(table.querySelectorAll('thead th, thead td'))
        .map(el => clean(el.innerText));
    const rows = Array.from(table.querySelectorAll('tbody tr')).map(tr => {
        const playerTh = tr.querySelector('th');
        const playerName = playerTh
            ? clean(playerTh.querySelector('a')?.getAttribute('aria-label') || '')
            : '';
        const position = playerTh
            ? clean(playerTh.querySelector('[class*="position"]')?.innerText || '')
            : '';
        const tdValues = Array.from(tr.querySelectorAll('td'))
            .map(td => clean(td.innerText));
        return [playerName, position, ...tdValues];
    }).filter(r => r[0]);
    const finalHeaders = [headers[0], 'POS', ...headers.slice(1)];
    return { headers: finalHeaders, rows };
}
"""

ROSTER_JS = """
() => {
  const clean = v => (v || '').replace(/\u00a0/g, ' ').replace(/\\s+/g, ' ').trim();
  return Array.from(document.querySelectorAll('h4')).flatMap(h4 => {
    const section = clean(h4.textContent);
    const container = h4.nextElementSibling;
    return Array.from(
      container?.querySelectorAll("div[class*='RosterPersonView__Container']") || []
    ).map(card => {
      const nameText = clean(card.querySelector("div[class*='RosterPersonView__PersonName']")?.textContent || '');
      const labels = Array.from(card.querySelectorAll("div[class*='InfoCompnents__InfoLabel']")).map(el => clean(el.textContent));
      const values = Array.from(card.querySelectorAll("div[class*='InfoCompnents__InfoData']")).map(el => clean(el.textContent));
      const info = {};
      labels.forEach((l, i) => { info[l] = values[i] || ''; });
      return { section, nameText, bt: info['B/T'] || '', hw: info['H/W'] || '' };
    });
  });
}
"""

# ── Helpers ────────────────────────────────────────────────────────────────────

def extract_table(page: Page) -> pd.DataFrame | None:
    data = page.evaluate(STATS_JS)
    if not data or not data.get("rows"):
        return None
    headers = data["headers"]
    n = len(headers)
    rows = [r[:n] + [""] * (n - len(r)) for r in data["rows"]]
    return pd.DataFrame(rows, columns=headers)


def select_dropdown(page: Page, index: int, option_text: str) -> bool:
    try:
        page.locator('[class*="css-2b097c-container"]').nth(index).click()
        page.wait_for_timeout(700)
        opt = page.locator(f'[class*="bui-dropdown__option"]:has-text("{option_text}")').first
        if not opt.is_visible(timeout=2000):
            page.keyboard.press("Escape")
            return False
        opt.click()
        page.wait_for_timeout(2500)
        return True
    except Exception:
        return False


def click_button(page: Page, text: str) -> bool:
    try:
        btn = page.locator(f"button:has-text('{text}')").first
        if btn.is_visible(timeout=2000):
            btn.click()
            page.wait_for_timeout(2000)
            return True
    except Exception:
        pass
    return False


def merge_std_exp(std: pd.DataFrame | None, exp: pd.DataFrame | None) -> pd.DataFrame | None:
    if std is None and exp is None:
        return None
    if std is None:
        return exp
    if exp is None:
        return std
    key_cols = [c for c in ["PLAYER", "POS", "TEAM"] if c in std.columns and c in exp.columns]
    exp_new = exp.drop(
        columns=[c for c in exp.columns if c in std.columns and c not in key_cols],
        errors="ignore",
    )
    return pd.merge(std, exp_new, on=key_cols, how="outer")


def parse_name_jersey(text: str) -> tuple[str, str | None]:
    text = re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()
    m = re.match(r"^(.*?)(?:\s+(\d{1,3}))?$", text)
    if not m:
        return text, None
    return re.sub(r"\s+", " ", m.group(1)).strip(), m.group(2)


def parse_bt(bt: str) -> tuple[str | None, str | None]:
    if "/" in bt:
        parts = bt.split("/")
        return parts[0].strip() or None, parts[1].strip() or None
    return None, None


def parse_hw(hw: str) -> tuple[float | None, float | None]:
    """Return (height_cm, weight_kg). Handles metric and imperial formats."""
    if not hw:
        return None, None
    height = weight = None

    # metric: "183 cm / 83.9 kg"
    m_cm = re.search(r"([\d.]+)\s*cm", hw, re.I)
    m_kg = re.search(r"([\d.]+)\s*kg", hw, re.I)
    if m_cm:
        height = float(m_cm.group(1))
    if m_kg:
        weight = float(m_kg.group(1))
    if height and weight:
        return height, weight

    # imperial compact: "6' 1\"/187"
    m_imp = re.search(r"(\d+)['\-]\s*(\d+)[\"']?\s*/\s*(\d+)", hw)
    if m_imp:
        feet, inches, lbs = int(m_imp.group(1)), int(m_imp.group(2)), int(m_imp.group(3))
        if not height:
            height = round(feet * 30.48 + inches * 2.54, 1)
        if not weight:
            weight = round(lbs * 0.453592, 1)
        return height, weight

    # imperial with "lbs" unit: "6' 1\" / 195 lbs"
    m_ft = re.search(r"(\d+)['\-]\s*(\d+)?[\"']?", hw)
    m_lbs = re.search(r"([\d.]+)\s*lbs", hw, re.I)
    if m_ft and not height:
        feet = int(m_ft.group(1))
        inches = int(m_ft.group(2)) if m_ft.group(2) else 0
        height = round(feet * 30.48 + inches * 2.54, 1)
    if m_lbs and not weight:
        weight = round(float(m_lbs.group(1)) * 0.453592, 1)

    return height, weight


# ── Stats scraping ─────────────────────────────────────────────────────────────

def scrape_year_stats(page: Page, team_filter: str, year: int) -> dict[str, pd.DataFrame]:
    """Scrape batting + pitching for one year. team_filter is the dropdown label."""
    results: dict[str, pd.DataFrame] = {}

    # Pitching: tab → team → year
    click_button(page, "Pitching")
    page.wait_for_timeout(2000)
    select_dropdown(page, 1, team_filter)
    select_dropdown(page, 0, str(year))

    click_button(page, "Standard")
    pit_std = extract_table(page)
    pit_exp = extract_table(page) if click_button(page, "Expanded") else None
    pit_df = merge_std_exp(pit_std, pit_exp)
    if pit_df is not None:
        results["Pitching"] = pit_df
        print(f"    Pitching: {len(pit_df)} rows")

    # Batting: Hitting tab → year
    click_button(page, "Hitting")
    page.wait_for_timeout(2000)
    select_dropdown(page, 0, str(year))

    click_button(page, "Standard")
    bat_std = extract_table(page)
    bat_exp = extract_table(page) if click_button(page, "Expanded") else None
    bat_df = merge_std_exp(bat_std, bat_exp)
    if bat_df is not None:
        results["Batting"] = bat_df
        print(f"    Batting:  {len(bat_df)} rows")

    return results


def build_player_rows(year: int, nationality: str, stats: dict[str, pd.DataFrame]) -> list[dict]:
    """One row per player, with BAT_*/PIT_* prefixed stat columns."""
    players: dict[str, dict] = {}

    for sheet_name, df in stats.items():
        prefix = "BAT_" if sheet_name == "Batting" else "PIT_"
        for _, row in df.iterrows():
            pname = row.get("PLAYER", "")
            if not pname:
                continue
            if pname not in players:
                players[pname] = {"player_name": pname, "POS": row.get("POS", ""), "TEAM": row.get("TEAM", "")}
            elif sheet_name == "Pitching" and not players[pname].get("POS"):
                players[pname]["POS"] = row.get("POS", "")
            for col, val in row.items():
                if col not in ("PLAYER", "POS", "TEAM"):
                    players[pname][prefix + col] = val

    rows = []
    for pname, data in players.items():
        pos = data.get("POS", "")
        rows.append({
            "season": year,
            "section": POS_TO_SECTION.get(pos, ""),
            "nationality": nationality,
            "player_name": pname,
            "player_name_zh": None,
            "team": data.get("TEAM", ""),
            "league": None,
            "jersey_no": None,
            "bats": None,
            "throws": None,
            "height": None,
            "weight": None,
            **{k: v for k, v in data.items() if k not in ("player_name", "POS", "TEAM")},
            "POS": pos,
        })
    return rows


def upsert_rows(all_rows: list[dict]) -> None:
    """Update existing Sheet1 rows or append new ones."""
    wb = load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}

    # Build index of existing rows
    row_index: dict[tuple, int] = {}
    for r in range(2, ws.max_row + 1):
        nat = ws.cell(r, col_idx.get("nationality", 1)).value
        season = ws.cell(r, col_idx.get("season", 1)).value
        name = ws.cell(r, col_idx.get("player_name", 1)).value
        if nat and season and name:
            row_index[(nat, season, name)] = r

    updated = appended = 0
    for row in all_rows:
        nat = row.get("nationality")
        season = row.get("season")
        name = row.get("player_name")
        target_row = row_index.get((nat, season, name))

        if target_row is None:
            target_row = ws.max_row + 1
            row_index[(nat, season, name)] = target_row

        for col_name, val in row.items():
            if col_name not in col_idx:
                new_col = ws.max_column + 1
                ws.cell(1, new_col).value = col_name
                col_idx[col_name] = new_col
                headers.append(col_name)
            if val is not None:
                existing = ws.cell(target_row, col_idx[col_name]).value
                if existing is None or str(existing).strip() == "":
                    ws.cell(target_row, col_idx[col_name]).value = val

        if target_row <= ws.max_row:
            updated += 1
        else:
            appended += 1

    wb.save(OUTPUT)
    print(f"  → updated={updated} appended={appended}")


def run_stats(teams_filter: list[str] | None, years_filter: list[int] | None) -> None:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1440, "height": 900})

        for slug, nat, team_filter, stat_years, _ in TEAMS:
            if teams_filter and slug not in teams_filter:
                continue
            years = [y for y in stat_years if not years_filter or y in years_filter]
            if not years:
                continue

            url = STATS_BASE_URL.format(slug=slug)
            print(f"\n{'='*55}\n  {nat}  {url}")
            page.goto(url, wait_until="domcontentloaded", timeout=120000)
            page.wait_for_timeout(8000)

            all_rows: list[dict] = []
            for year in years:
                print(f"  [{year}]")
                stats = scrape_year_stats(page, team_filter, year)
                all_rows.extend(build_player_rows(year, nat, stats))

            upsert_rows(all_rows)

        browser.close()
    print("\nStats scraping done.")


# ── Roster info filling ────────────────────────────────────────────────────────

def scrape_roster_page(page: Page, slug: str, year: int) -> list[dict]:
    url = ROSTER_BASE_URL.format(slug=slug, year=year)
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(4000)
    if NO_DATA in page.locator("body").inner_text():
        return []
    raw = page.evaluate(ROSTER_JS)
    rows = []
    for item in raw:
        name, jersey = parse_name_jersey(item["nameText"])
        if not name:
            continue
        bats, throws = parse_bt(item["bt"])
        height, weight = parse_hw(item["hw"])
        rows.append({
            "player_name": name,
            "jersey_no": int(jersey) if jersey else None,
            "section": item["section"],
            "bats": bats,
            "throws": throws,
            "height": height,
            "weight": weight,
        })
    return rows


def run_roster(teams_filter: list[str] | None, years_filter: list[int] | None) -> None:
    all_data: dict[tuple[str, int], list[dict]] = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1440, "height": 2200})

        for slug, nat, _, _, roster_years in TEAMS:
            if teams_filter and slug not in teams_filter:
                continue
            if not roster_years:
                continue
            years = [y for y in roster_years if not years_filter or y in years_filter]
            for year in years:
                print(f"  {nat} {year} ...", end=" ", flush=True)
                players = scrape_roster_page(page, slug, year)
                if players:
                    all_data[(nat, year)] = players
                    print(f"{len(players)} players")
                else:
                    print("no data")

        browser.close()

    # Write to Sheet1
    wb = load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    row_index: dict[tuple, int] = {}
    for r in range(2, ws.max_row + 1):
        nat = ws.cell(r, col["nationality"]).value
        season = ws.cell(r, col["season"]).value
        name = ws.cell(r, col["player_name"]).value
        if nat and season and name:
            row_index[(nat, season, name)] = r

    updated = unmatched = 0
    for (nat, year), players in all_data.items():
        for p in players:
            r = row_index.get((nat, year, p["player_name"]))
            if r is None:
                unmatched += 1
                continue
            for field in ("jersey_no", "section", "bats", "throws", "height", "weight"):
                val = p.get(field)
                if val is not None and col.get(field):
                    ws.cell(r, col[field]).value = val
            updated += 1

    wb.save(OUTPUT)
    print(f"\nRoster info: updated={updated} unmatched={unmatched}")
    print("Roster filling done.")


# ── CLI ────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="WBC scraper")
    parser.add_argument("command", choices=["stats", "roster", "all"],
                        help="stats=scrape batting/pitching, roster=fill player info, all=both")
    parser.add_argument("--teams", nargs="+", metavar="SLUG",
                        help="Filter by team slug(s), e.g. japan korea")
    parser.add_argument("--years", nargs="+", type=int, metavar="YEAR",
                        help="Filter by year(s), e.g. 2026 2023")
    args = parser.parse_args()

    if args.command in ("stats", "all"):
        print("=== Scraping stats ===")
        run_stats(args.teams, args.years)

    if args.command in ("roster", "all"):
        print("\n=== Filling roster info ===")
        run_roster(args.teams, args.years)


if __name__ == "__main__":
    main()
