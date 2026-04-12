"""
Fill player_name_zh by querying the Wikipedia API for native-script names.
For Japanese players: kanji names (大谷翔平)
For Korean players: Hangul names (오타니 쇼헤이) — but we also try Chinese name if available
"""
from __future__ import annotations
import re
import time
import json
from pathlib import Path
import urllib.request
import urllib.parse
from openpyxl import load_workbook

OUTPUT = Path(__file__).parent.parent / "wbc_roster.xlsx"

# Hard-coded overrides for names Wikipedia can't easily find
MANUAL: dict[str, str] = {
    # Japan famous players (kanji = Chinese)
    "Shohei Ohtani":        "大谷翔平",
    "Yu Darvish":           "達比修有",
    "Masahiro Tanaka":      "田中將大",
    "Ichiro Suzuki":        "鈴木一朗",
    "Hideki Matsui":        "松井秀喜",
    "Kenji Johjima":        "城島健司",
    "Daisuke Matsuzaka":    "松坂大輔",
    "Koji Uehara":          "上原浩治",
    "Hiroki Kuroda":        "黑田博樹",
    # Players on Japan's WBC team but not Japanese nationals → skip CJK name
    "Lars Nootbaar":        "",
    "Tommy Edman":          "",
    # Korea famous
    "Chan Ho Park":         "朴贊浩",
    "Hyun-Jin Ryu":         "柳賢振",
    "Ha-Seong Kim":         "金河成",
    "ByungHo Park":         "朴炳鎬",
    "Hyeseong Kim":         "金慧成",
    "Jung Hoo Lee":         "李政厚",
    "Baek Ho Kang":         "姜白虎",
}

# Names to skip (non-nationals playing for Japan/Korea team)
SKIP_NAMES: set[str] = {"Lars Nootbaar", "Tommy Edman"}


HEADERS = {"User-Agent": "WBC-Stats-Bot/1.0 (baseball stats research)"}


def _api_get(url: str) -> dict:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=12) as resp:
        return json.loads(resp.read())


def _name_matches(en_name: str, page_title: str) -> bool:
    """Check page title is plausibly about this player (not a false positive)."""
    en_parts = set(en_name.lower().replace("-", " ").split())
    title_parts = set(page_title.lower().replace("-", " ").split())
    # Accept if at least 2 words match, or page title contains English words matching name
    if len(en_parts & title_parts) >= min(2, len(en_parts)):
        return True
    # Also accept if title is fully CJK but page snippet mentions the English name
    return False


def wikipedia_search_native_name(en_name: str, nationality: str) -> str:
    """Query Wikipedia API and extract CJK name from infobox wikitext."""
    search_url = (
        "https://en.wikipedia.org/w/api.php?action=query&list=search"
        f"&srsearch={urllib.parse.quote(en_name + ' baseball')}"
        "&srlimit=3&format=json"
    )
    try:
        data = _api_get(search_url)
        results = data.get("query", {}).get("search", [])
        # Pick the best matching result
        page_title = ""
        for r in results:
            if _name_matches(en_name, r["title"]):
                page_title = r["title"]
                break
        if not page_title:
            return ""

        parse_url = (
            "https://en.wikipedia.org/w/api.php?action=parse"
            f"&page={urllib.parse.quote(page_title)}"
            "&prop=wikitext&section=0&format=json"
        )
        data2 = _api_get(parse_url)
        wikitext = data2.get("parse", {}).get("wikitext", {}).get("*", "")

        # Verify page is about a baseball player (not a false positive)
        baseball_keywords = ("baseball", "pitcher", "outfield", "infield", "catcher",
                             "NPB", "KBO", "MLB", "shortstop", "designated hitter")
        if not any(kw.lower() in wikitext[:2000].lower() for kw in baseball_keywords):
            return ""

        # Japanese players: kanji (same characters work in Chinese)
        if nationality == "日本":
            for pattern in [
                r"\|\s*native_name\s*=\s*([^\n|]+)",
                r"\|\s*kanji\s*=\s*([^\n|]+)",
            ]:
                m = re.search(pattern, wikitext)
                if m:
                    cjk = re.sub(r"[^\u3000-\u9fff]", "", m.group(1))
                    if 2 <= len(cjk) <= 8:
                        return cjk
            # Fallback: find consecutive kanji sequences to build full name
            # "鈴木 誠也" → positions of each sequence, concatenate nearby ones
            it = list(re.finditer(r"[\u4e00-\u9fff\u3400-\u4dbf]{1,}", wikitext[:3000]))
            if it:
                # Try to find a pair/single that looks like a name (3-5 CJK chars total)
                for idx, m in enumerate(it):
                    seq = m.group()
                    # Check if next sequence is close (within 3 chars gap = space only)
                    if idx + 1 < len(it):
                        next_m = it[idx + 1]
                        gap = next_m.start() - m.end()
                        combined = seq + next_m.group()
                        if gap <= 3 and 3 <= len(combined) <= 5:
                            return combined
                    if 2 <= len(seq) <= 4:
                        return seq

        # Korean players: try hanja field, else skip (Hangul ≠ Chinese)
        elif nationality == "韓國":
            for pattern in [
                r"\|\s*(?:hanja|hanja_name)\s*=\s*([^\n|]+)",
            ]:
                m = re.search(pattern, wikitext)
                if m:
                    cjk = re.sub(r"[^\u3400-\u9fff]", "", m.group(1))
                    if 2 <= len(cjk) <= 8:
                        return cjk

    except Exception:
        pass
    return ""


def main() -> None:
    wb = load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    # Collect rows needing Chinese names: (row, en_name, nationality)
    needs = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col["player_name_zh"]).value:
            continue
        nat = ws.cell(r, col["nationality"]).value
        if nat not in ("日本", "韓國"):
            continue
        name = ws.cell(r, col["player_name"]).value
        if name:
            needs.append((r, name, nat))

    # Deduplicate: (name, nat) → zh
    unique: dict[tuple, str] = {}
    for _, name, nat in needs:
        key = (name, nat)
        if key not in unique:
            unique[key] = MANUAL.get(name, "")

    to_lookup = [(name, nat) for (name, nat), zh in unique.items() if not zh]
    print(f"Manual overrides: {len(unique) - len(to_lookup)} | API lookups: {len(to_lookup)}")

    for i, (name, nat) in enumerate(to_lookup):
        zh = wikipedia_search_native_name(name, nat)
        unique[(name, nat)] = zh
        status = zh if zh else "(not found)"
        print(f"  [{i+1}/{len(to_lookup)}] {name} ({nat}) → {status}")
        time.sleep(0.35)

    # Write back
    filled = 0
    for r, name, nat in needs:
        zh = unique.get((name, nat), "")
        if zh:
            ws.cell(r, col["player_name_zh"]).value = zh
            filled += 1

    wb.save(OUTPUT)
    print(f"\nFilled {filled} / {len(needs)} rows with Chinese names.")


if __name__ == "__main__":
    main()
