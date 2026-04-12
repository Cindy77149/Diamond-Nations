"""
Calculate player ability ratings based on WBC stats.
Uses percentile ranking within the full dataset, scaled to 40-99.
Adds columns: RTG_CONTACT, RTG_POWER, RTG_EYE, RTG_SPEED, RTG_BAT,
              RTG_CTRL, RTG_K, RTG_ERA_Q, RTG_PIT, RTG_OVR
"""
from __future__ import annotations
import numpy as np
from pathlib import Path
from openpyxl import load_workbook

OUTPUT = Path(__file__).parent.parent / "wbc_roster.xlsx"

MIN_PA   = 5      # 打者最低打席門檻
MIN_IP   = 2.0    # 投手最低投球局數
DEFAULT  = 68     # 門檻以下/無數據預設分（有參賽但樣本不足）
LO, HI   = 58, 99 # 分數範圍：底部從 58 開始，避免極端低分失真

RTG_COLS = [
    "RTG_CONTACT", "RTG_POWER", "RTG_EYE", "RTG_SPEED", "RTG_BAT",
    "RTG_CTRL",    "RTG_K",     "RTG_ERA_Q", "RTG_PIT",  "RTG_OVR",
]


# ── helpers ──────────────────────────────────────────────────────────────────

def to_f(v) -> float | None:
    try:
        f = float(v)
        return None if np.isnan(f) else f
    except (TypeError, ValueError):
        return None


def pct_score(value: float, pool: list[float], higher_better: bool = True) -> int:
    """Convert a value to 40-99 using percentile rank in pool."""
    if not pool:
        return DEFAULT
    pct = sum(1 for x in pool if x <= value) / len(pool) * 100
    if not higher_better:
        pct = 100 - pct
    return round(LO + (HI - LO) * pct / 100)


def weighted(*pairs) -> float:
    """Weighted average of (value, weight) pairs, ignoring None."""
    total_w = total_v = 0.0
    for v, w in pairs:
        if v is not None:
            total_v += v * w
            total_w += w
    return (total_v / total_w) if total_w else None


def clamp(v) -> int | None:
    if v is None:
        return None
    return max(LO, min(HI, round(v)))


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    # Add rating columns if not present
    next_col = ws.max_column + 1
    for rc in RTG_COLS:
        if rc not in col:
            ws.cell(1, next_col).value = rc
            col[rc] = next_col
            next_col += 1

    # ── Read all rows ─────────────────────────────────────────────────────────
    rows = []
    for r in range(2, ws.max_row + 1):
        def g(c):
            return to_f(ws.cell(r, col[c]).value) if c in col else None

        pa  = g("BAT_PA")
        # 若 PA 缺失，用 AB+BB+HBP+SAC+SF 補算
        if pa is None:
            ab  = to_f(g("BAT_AB"))  or 0
            bb  = to_f(g("BAT_BB"))  or 0
            hbp = to_f(g("BAT_HBP")) or 0
            sac = to_f(g("BAT_SAC")) or 0
            sf  = to_f(g("BAT_SF"))  or 0
            est = ab + bb + hbp + sac + sf
            if est > 0:
                pa = est
        ip  = g("PIT_IP")
        pos = ws.cell(r, col["POS"]).value if "POS" in col else None

        rows.append({
            "row":       r,
            "pa":        pa,
            "ip":        ip,
            "pos":       pos,
            # batting
            "avg":       g("BAT_AVG"),
            "obp":       g("BAT_OBP"),
            "slg":       g("BAT_SLG"),
            "ops":       g("BAT_OPS"),
            "iso":       g("BAT_ISO"),
            "bb_pct":    g("BAT_BB%"),
            "k_pct":     g("BAT_K%"),
            "sb":        g("BAT_SB"),
            "cs":        g("BAT_CS"),
            "hr":        g("BAT_HR"),
            # pitching
            "era":       g("PIT_ERA"),
            "whip":      g("PIT_WHIP"),
            "so9":       g("PIT_SO/9"),
            "bb9":       g("PIT_BB/9"),
            "kbb":       g("PIT_K/BB"),
            "pit_avg":   g("PIT_AVG"),
            "sv":        g("PIT_SV"),
            "hld":       g("PIT_HLD"),
            "qs":        g("PIT_QS"),
            "gs":        g("PIT_GS"),
        })

    # ── Build pools (qualified players only) ─────────────────────────────────
    def pool(key, min_key, min_val):
        return [r[key] for r in rows
                if r[key] is not None and r[min_key] is not None and r[min_key] >= min_val]

    p_avg    = pool("avg",     "pa",  MIN_PA)
    p_obp    = pool("obp",     "pa",  MIN_PA)
    p_slg    = pool("slg",     "pa",  MIN_PA)
    p_ops    = pool("ops",     "pa",  MIN_PA)
    p_iso    = pool("iso",     "pa",  MIN_PA)
    p_bb_pct = pool("bb_pct",  "pa",  MIN_PA)
    p_k_pct  = pool("k_pct",   "pa",  MIN_PA)
    p_sb     = pool("sb",      "pa",  MIN_PA)
    p_hr     = pool("hr",      "pa",  MIN_PA)

    p_era    = pool("era",     "ip",  MIN_IP)
    p_whip   = pool("whip",    "ip",  MIN_IP)
    p_so9    = pool("so9",     "ip",  MIN_IP)
    p_bb9    = pool("bb9",     "ip",  MIN_IP)
    p_kbb    = pool("kbb",     "ip",  MIN_IP)
    p_pavg   = pool("pit_avg", "ip",  MIN_IP)

    # ── Score each row ────────────────────────────────────────────────────────
    for d in rows:
        r    = d["row"]
        pa   = d["pa"]   or 0
        ip   = d["ip"]   or 0
        bat_qualified = pa  >= MIN_PA
        pit_qualified = ip  >= MIN_IP

        # ── BATTING ──────────────────────────────────────────────────────────
        if bat_qualified:
            contact_s = clamp(weighted(
                (pct_score(d["avg"],   p_avg,   True)  if d["avg"]   is not None else None, 0.5),
                (pct_score(d["k_pct"], p_k_pct, False) if d["k_pct"] is not None else None, 0.5),
            ))
            power_s = clamp(weighted(
                (pct_score(d["iso"],   p_iso,  True) if d["iso"]  is not None else None, 0.4),
                (pct_score(d["slg"],   p_slg,  True) if d["slg"]  is not None else None, 0.4),
                (pct_score(d["hr"],    p_hr,   True) if d["hr"]   is not None else None, 0.2),
            ))
            eye_s = clamp(weighted(
                (pct_score(d["bb_pct"], p_bb_pct, True) if d["bb_pct"] is not None else None, 0.5),
                (pct_score(d["obp"],    p_obp,    True) if d["obp"]    is not None else None, 0.5),
            ))
            # speed: SB percentile; adjust down for caught stealing
            if d["sb"] is not None:
                sb_s = pct_score(d["sb"], p_sb, True)
                if d["sb"] + (d["cs"] or 0) > 0:
                    cs_pen = (d["cs"] or 0) / (d["sb"] + (d["cs"] or 0))
                    sb_s = clamp(sb_s * (1 - cs_pen * 0.3))
                speed_s = clamp(sb_s)
            else:
                speed_s = DEFAULT

            bat_s = clamp(weighted(
                (pct_score(d["ops"], p_ops, True) if d["ops"] is not None else None, 0.5),
                (contact_s, 0.2),
                (power_s,   0.2),
                (eye_s,     0.1),
            ))
        else:
            contact_s = power_s = eye_s = speed_s = bat_s = DEFAULT if pa > 0 else None

        # ── PITCHING ─────────────────────────────────────────────────────────
        if pit_qualified:
            ctrl_s = clamp(weighted(
                (pct_score(d["whip"], p_whip, False) if d["whip"] is not None else None, 0.5),
                (pct_score(d["bb9"],  p_bb9,  False) if d["bb9"]  is not None else None, 0.5),
            ))
            k_s = clamp(weighted(
                (pct_score(d["so9"], p_so9, True) if d["so9"] is not None else None, 0.5),
                (pct_score(d["kbb"], p_kbb, True) if d["kbb"] is not None else None, 0.5),
            ))
            era_q_s = clamp(weighted(
                (pct_score(d["era"],     p_era,  False) if d["era"]     is not None else None, 0.6),
                (pct_score(d["pit_avg"], p_pavg, False) if d["pit_avg"] is not None else None, 0.4),
            ))
            pit_s = clamp(weighted(
                (era_q_s, 0.4),
                (ctrl_s,  0.3),
                (k_s,     0.3),
            ))
        else:
            ctrl_s = k_s = era_q_s = pit_s = DEFAULT if ip > 0 else None

        # ── OVERALL ──────────────────────────────────────────────────────────
        pos_str = str(d["pos"] or "")
        is_pitcher = pos_str in ("SP", "RP", "CP", "SP/RP", "P")

        if is_pitcher and pit_s is not None:
            ovr = pit_s
        elif not is_pitcher and bat_s is not None:
            ovr = bat_s
        elif pit_s is not None:
            ovr = pit_s
        elif bat_s is not None:
            ovr = bat_s
        else:
            ovr = DEFAULT

        # ── Write ────────────────────────────────────────────────────────────
        def w(c, v):
            ws.cell(r, col[c]).value = v if v is not None else DEFAULT

        w("RTG_CONTACT", contact_s)
        w("RTG_POWER",   power_s)
        w("RTG_EYE",     eye_s)
        w("RTG_SPEED",   speed_s)
        w("RTG_BAT",     bat_s)
        w("RTG_CTRL",    ctrl_s)
        w("RTG_K",       k_s)
        w("RTG_ERA_Q",   era_q_s)
        w("RTG_PIT",     pit_s)
        w("RTG_OVR",     ovr)

    wb.save(OUTPUT)

    # ── Summary ───────────────────────────────────────────────────────────────
    ovr_vals = [ws.cell(r, col["RTG_OVR"]).value for r in range(2, ws.max_row + 1)]
    ovr_vals = [v for v in ovr_vals if v is not None]
    print(f"Rows processed: {len(rows)}")
    print(f"OVR  avg={sum(ovr_vals)/len(ovr_vals):.1f}  "
          f"min={min(ovr_vals)}  max={max(ovr_vals)}")

    # Top 10
    top = sorted(
        [(ws.cell(r, col["RTG_OVR"]).value,
          ws.cell(r, col["player_name"]).value,
          ws.cell(r, col["player_name_zh"]).value,
          ws.cell(r, col["nationality"]).value,
          ws.cell(r, col["season"]).value,
          ws.cell(r, col["POS"]).value)
         for r in range(2, ws.max_row + 1)],
        reverse=True
    )[:15]
    print("\nTop 15 OVR:")
    for ovr, en, zh, nat, sea, pos in top:
        zh = zh or ""
        print(f"  {ovr:3d}  {en:25s} {zh:8s} {nat} {sea} {pos}")


if __name__ == "__main__":
    main()
