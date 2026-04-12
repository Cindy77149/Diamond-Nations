"""
gen_player_js.py
Read wbc_roster.xlsx → generate players/<nat>.js
"""
from __future__ import annotations
import math
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

XLSX       = Path(__file__).parent.parent / "wbc_roster.xlsx"
OUTPUT     = Path(__file__).parent.parent / "players-all.js"

PITCHER_POS = {"SP", "RP", "CP", "SP/RP", "P"}

NAT_TO_FLAG = {
    "台灣": "🇹🇼", "日本": "🇯🇵", "韓國": "🇰🇷", "美國": "🇺🇸",
    "多明尼加": "🇩🇴", "委內瑞拉": "🇻🇪", "古巴": "🇨🇺", "波多黎各": "🇵🇷",
}
NAT_TO_FILE = {
    "台灣": "tw", "日本": "jp", "韓國": "kr", "美國": "us",
    "多明尼加": "do", "委內瑞拉": "ve", "古巴": "cu", "波多黎各": "pr",
}
NAT_DISPLAY = {
    "台灣": "🇹🇼 台灣", "日本": "🇯🇵 日本", "韓國": "🇰🇷 韓國", "美國": "🇺🇸 美國",
    "多明尼加": "🇩🇴 多明尼加", "委內瑞拉": "🇻🇪 委內瑞拉", "古巴": "🇨🇺 古巴", "波多黎各": "🇵🇷 波多黎各",
}

POS_AV = {
    "C": "🛡️", "1B": "💪", "2B": "⚡", "3B": "🔥", "SS": "⚡",
    "LF": "🏃", "CF": "💨", "RF": "💥", "OF": "🏃", "DH": "💣",
    "SP": "🎯", "RP": "⚡", "CP": "🔒", "SP/RP": "🌟", "P": "⚾",
    "INF": "🧤", "UT": "🔧",
}

# Positional defense baseline (batters only)
DEF_BY_POS = {
    "C": 85, "SS": 82, "2B": 80, "CF": 78, "3B": 76,
    "LF": 73, "RF": 73, "1B": 68, "OF": 74, "DH": 60,
    "INF": 76, "UT": 73,
}

# Manual defense overrides for elite defenders (keyed by English player name)
# These ensure famous defensive players get appropriate defense stats
# regardless of their batting OVR
DEF_OVERRIDES: dict[str, int] = {
    # ── Catchers ──────────────────────────────────────────────────────────────
    "Yadier Molina":        96,   # 歷史最強守備捕手之一
    "Ivan Rodriguez":       97,   # 名人堂・13 座 MLB 黃金手套・史上最強守備捕手
    "Martín Maldonado":     90,   # 以守備/接球見長
    "J.T. Realmuto":        88,
    "Salvador Perez":       86,
    # ── Shortstops ────────────────────────────────────────────────────────────
    "Omar Vizquel":         97,   # 史上最強守備游擊手之一・11 座黃金手套
    "Hayato Sakamoto":      94,   # 日本最強守備游擊手・黃金手套 10 度
    "Shinya Miyamoto":      90,   # 日本傳奇守備游擊手
    "Ha-Seong Kim":         92,   # 黃金手套・MLB 頂級守備
    "Francisco Lindor":     90,   # 黃金手套游擊手
    "Alcides Escobar":      88,   # 守備型游擊手
    "Javier Báez":          88,   # 守備範圍超強
    "Brandon Crawford":     88,   # 黃金手套 SS
    "Munenori Kawasaki":    86,   # 以守備著稱
    "Sosuke Genda":         92,   # 日本・7 連座 NPB 黃金手套游擊手
    "Chin-Lung Hu":         88,   # 台灣傳奇守備游擊手・曾被道奇寄予厚望
    "Kun-Yu Chiang":        92,   # 台灣・6 連座 CPBL 黃金手套・史上首位内野手
    "Yu Chang":             90,   # 台灣・部長・多守備位工具型・MLB 游擊守備扎實
    "Ji Hwan Oh":           84,   # 韓國・2022 KBO 黃金手套游擊手
    "Elvis Andrus":         82,   # 委內瑞拉・守備型游擊手
    "Cesar Izturis":        88,   # 委內瑞拉・黃金手套 SS
    "Erisbel Arruebarrena": 90,   # 古巴・頂尖守備游擊
    "José Reyes":           84,   # 多明尼加・速度+守備型 SS
    "Jimmy Rollins":        85,   # 美國・黃金手套 SS
    "Carlos Correa":        90,   # 波多黎各・黃金手套 SS/3B
    "Trea Turner":          92,   # 美國・守備範圍頂尖・Gold Glove 等級游擊手
    # ── Second Basemen ────────────────────────────────────────────────────────
    "Robinson Canó":        85,   # 黃金手套二壘手
    "Placido Polanco":      86,
    "Hyeseong Kim":         92,   # 韓國・4 座 KBO 黃金手套 + 2 座 KBO 守備獎・首位游/二壘都拿獎的球員
    "Ryosuke Kikuchi":      95,   # 日本・10 連座 NPB 黃金手套二壘手・NPB 史上最強守備二壘手
    "Takumu Nakano":        86,   # 日本・2023、2025 NPB 黃金手套・終結菊池 10 連霸
    "Chase Utley":          85,   # 美國・黃金手套 2B
    "Dustin Pedroia":       85,   # 美國・黃金手套 2B
    "Brandon Phillips":     85,   # 美國・黃金手套 2B
    "Andrés Giménez":       92,   # 委內瑞拉・3 連座 MLB 黃金手套 + AL 鉑金手套
    "Ian Kinsler":          84,   # 美國・2018 MLB 黃金手套 2B
    "Alfonso Soriano":      70,   # 多明尼加・守備差被迫改打 LF・2B 守備評分長期墊底
    # ── Center Fielders ───────────────────────────────────────────────────────
    "Adam Jones":           88,   # 黃金手套中外野
    "Ender Inciarte":       90,   # 黃金手套中外野・頂級守備
    "Pete Crow-Armstrong":  85,   # 以守備出名
    "Byron Buxton":         90,   # 美國・2023 MLB 黃金手套 CF・頂尖守備範圍
    "Ken Griffey Jr.":      92,   # 美國・10 座 MLB 黃金手套 CF・史上最佳守備中外野之一
    "Ichiro Suzuki":        95,   # 10 座黃金手套・史上最強外野臂力之一
    "Kosuke Fukudome":      86,   # 日本・5 座 NPB 黃金手套 CF
    "Norichika Aoki":       84,   # 日本・3 座 NPB 黃金手套・守備範圍優秀
    "Seiya Suzuki":         84,   # 日本・5 座 NPB 中央聯盟黃金手套 CF
    "Szu-Chi Chou":         75,   # 台灣・中職安打王・生涯外野手・守備稱職非金手套
    "Chieh-Hsien Chen":     86,   # 台灣・陳傑憲・4 連座 CPBL 黃金手套外野
    "Shogo Akiyama":        82,   # NPB 守備佳的 CF
    "Lars Nootbaar":        82,   # 以守備見長
    "Jung Hoo Lee":         88,   # KBO 黃金手套 CF
    "Leonys Martin":        86,   # 古巴・速度型守備 CF
    "Luis Robert Jr.":      90,   # 古巴・2020 MLB 黃金手套 CF（新人年）・頂尖臂力+守備
    "Endy Chavez":          84,   # 委內瑞拉・守備型外野
    "Gerardo Parra":        86,   # 委內瑞拉・3 座黃金手套
    "Mike Trout":           84,   # 全方位・守備亦優秀
    "Che-Hsuan Lin":        86,   # 台灣・NPB 黃金手套 CF
    "Chien-Ming Chang":     87,   # 台灣・張建銘・守備型右外野手
    # ── Right Fielders ────────────────────────────────────────────────────────
    "Mookie Betts":         90,   # 美國・6 座黃金手套
    "Ah-seop Son":          85,   # 韓國・KBO 黃金手套 RF
    "Alex Rios":            84,   # 波多黎各・守備優秀 RF
    "Fernando Tatis Jr.":   88,   # 多明尼加・臂力+守備頂尛 RF
    "Nelson Cruz":          68,   # 多明尼加・生涯 DRS 長期負值・著名守備差外野
    # ── Third Basemen ─────────────────────────────────────────────────────────
    "Nolan Arenado":        97,   # 史上最強守備三壘手・10 連座 MLB 黃金手套
    "Manny Machado":        90,   # 黃金手套三壘/游擊
    "Adrian Beltré":        95,   # 多明尼加・傳奇守備三壘手・5 座黃金手套
    "Nobuhiro Matsuda":     88,   # 日本・NPB 多次黃金手套 3B
    "Evan Longoria":        88,   # 美國・3 座 MLB 黃金手套三壘手
    "Miguel Cabrera":       70,   # 委內瑞拉・3B 守備平庸・生涯 DRS 持續負值・後期移至 1B
    # ── First Basemen ─────────────────────────────────────────────────────────
    "Albert Pujols":        84,   # 一壘守備頂尖
    "Derrek Lee":           88,   # 美國・3 座黃金手套 1B
    "Eric Hosmer":          84,   # 美國・4 座黃金手套 1B
    "Mark Teixeira":        82,   # 美國・5 座黃金手套 1B
    # ── Catchers (extra) ──────────────────────────────────────────────────────
    "Roberto Pérez":        88,   # 波多黎各・頂級接球/守備捕手
    "Kyle Higashioka":      82,   # 美國・以守備見長的後備捕手
    "Ariel Pestano":        86,   # 古巴傳奇守備捕手
    "Buster Posey":         85,   # 美國・守備優秀
    "Jason Varitek":        82,   # 美國・指揮型捕手・守備佳
    "Adam Dunn":            60,   # 美國・MLB 史上最差守備外野之一・生涯 DRS 長期 -20 以下
    # ── Poor defensive catchers ───────────────────────────────────────────────
    "Carlos Santana":       70,   # 多明尼加・守備差出名・生涯後期被迫移至 1B
    "Ronny Paulino":        76,   # 多明尼加・備用捕手・守備無特別之處
    "Austin Wells":         80,   # 多明尼加・發展中捕手・2026 守備仍待成熟
}

# Manual mental overrides for clutch/big-game players (keyed by English player name)
# Based on documented WBC key moments, career clutch reputation, and international performance
MEN_OVERRIDES: dict[str | tuple, int] = {
    # ── 🇯🇵 日本 ────────────────────────────────────────────────────────────────
    "Shohei Ohtani":        99,   # 2023 MVP・賽前演講・決賽三振 Trout・史上最頂級大賽心理
    "Ichiro Suzuki":        98,   # 2009 決賽第 10 局 2 好 2 壞 2 出局・低潮後送出制勝安打
    "Daisuke Matsuzaka":    97,   # WBC 史上唯一 2 連霸 MVP・賽場上永遠泰然自若
    "Munetaka Murakami":    96,   # 2023 半決賽 4 打數 0 安後，第 9 局送出致勝二壘安打
    "Masahiro Tanaka":      95,   # 24 勝 0 敗 2013 賽季・日本大賽 Game 7 投 160 球後隔日救援
    "Koji Uehara":          95,   # WBC 全程 0 四壞・2013 世界大賽終結者・史上 K/BB 最頂尖
    "Masataka Yoshida":     94,   # 2023 WBC RBI 紀錄（13）・7 場全勤・穩定超強
    "Roki Sasaki":          93,   # 2023 WBC 開幕戰 6 局完封韓國・決賽回板壓制美國・史上最強 WBC 先發表現之一
    "Yu Darvish":           93,   # 2009 決賽被追平後在延長局守住冠軍・抗壓復原力強
    "Hayato Sakamoto":      92,   # 代表隊常勝精神核心・多屆 WBC 中軸
    "Sosuke Genda":         90,   # 2023 WBC 守備精神支柱
    # ── 🇰🇷 韓國 ────────────────────────────────────────────────────────────────
    "Seunghwan Oh":         97,   # 「石佛」・KBO 史上最強後援・神情不動如山・400+ 救援
    "Lee Seung-yeop":       94,   # 2006 WBC 5 轟平紀錄・10 打點・韓國全壘打王
    "Hyun-Jin Ryu":         92,   # 19 歲完封加拿大・多屆 WBC 韓國王牌
    "Hyun Soo Kim":         89,   # 2009 決賽 9 局最高壓四壞球・All-WBC 球員
    "Bum Ho Lee":           89,   # 2009 決賽第 9 局 2 出局追平安打・最高壓一棒
    "Shin-Soo Choo":        88,   # 半決賽 3 分砲・大賽穩定老將
    "Ha-Seong Kim":         88,   # 韓國 WBC 守攻雙全中堅
    "Jung Hoo Lee":         88,   # KBO 大賽長期穩定
    # ── 🇹🇼 台灣 ────────────────────────────────────────────────────────────────
    "Chien-Ming Wang":      92,   # 2013 WBC 連投 12 局封鎖・低谷後王者回歸・最強台灣大賽時刻
    "Yu Chang":             90,   # 2023 Pool A MVP・8 打點・大滿貫・超越日常 MLB 表現
    "Chieh-Hsien Chen":     89,   # Premier 12 MVP・CPBL 大賽長期核心・心理穩定
    "Kungkuan Giljegiljaw": 87,   # 2023 WBC 第 8 局 3 分砲・關鍵時刻一棒
    # ── 🇺🇸 美國 ────────────────────────────────────────────────────────────────
    "Marcus Stroman":       93,   # 2017 WBC MVP・冠軍賽近乎完全比賽・拼勁型大心臟
    "Trea Turner":          91,   # 2023 WBC 5 轟平紀錄・11 打點・帶動美國攻勢
    "Adam Jones":           88,   # 2017 關鍵守備飛撲搶走 Machado 轟・改變冠軍走向
    "Ian Kinsler":          87,   # 2017 冠軍賽首局 2 分砲・定調美國獲勝基調
    "Alex Rodriguez":       72,   # 多次季後賽大賽關鍵時刻消失・反心理代表人物
    "Jeurys Familia":       78,   # 2016 家暴停賽・非 clutch 型終結者
    "José De León":         80,   # 兩次 Tommy John 手術・生涯不足 200 局・傷病型投手
    # ── 🇩🇴 多明尼加 ────────────────────────────────────────────────────────────
    "Wander Franco":        68,   # 場外問題・人格不穩定因素・大賽表現反覆
    "Robinson Canó":        95,   # 2013 三獎 MVP 全制霸・WBC 安打紀錄・無敗冠軍核心
    # ── 年份專屬覆蓋（tuple key：名聲建立前的年份）─────────────────────────────
    ("Bum Ho Lee", 2006):   76,   # 2006 時尚未有 2009 決賽關鍵安打・回歸公式值
    "Fernando Rodney":      94,   # 2013 全程 7 救援 7 成功・0 分・史上最完美 WBC 終結者
    "Nelson Cruz":          92,   # 2017 對 Miller 0 好 2 壞送出 3 分砲・最不可思議的大賽一棒
    "Edwin Encarnación":    90,   # 2013 決賽致勝 2 分二壘打・棄走 Canó 改投他・結果付出代價
    "José Reyes":           88,   # 2013 無敗冠軍全程穩定開路先鋒
    "Fernando Tatis Jr.":   88,   # 全方位大賽球員・天賦級競爭心
    "Adrian Beltré":        90,   # 名人堂三壘手・長期大賽穩定代表
    # ── 🇻🇪 委內瑞拉 ────────────────────────────────────────────────────────────
    "Félix Hernández":      92,   # 2009 親口說是生涯最重要一場・8.2 局封鎖晉級
    "Francisco Rodríguez":  91,   # K-Rod・MLB 單季 62 救援紀錄・WBC 3 次救援成功
    "Miguel Cabrera":       91,   # 三冠王・史上最全能的打者之一・大賽精神主軸
    "Bobby Abreu":          88,   # 職業型打席・WBC 對手點名最難纏
    "Omar Vizquel":         90,   # 游擊傳奇・大賽精神領袖
    "Andrés Giménez":       88,   # 三連座 GG・持續在大賽交出成績
    # ── 🇨🇺 古巴 ────────────────────────────────────────────────────────────────
    "Pedro Luis Lazo":      93,   # 2006 半決賽壓制全明星打線晉決賽・古巴終結者傳奇
    "Frederich Cepeda":     93,   # 2009 WBC .500/.538/.958・多屆 WBC 最穩定古巴攻擊核心
    "Yuli Gurriel":         90,   # 2006 WBC All-Tournament・古巴最頂尖野手・後赴 MLB
    "Ariel Pestano":        89,   # 古巴傳奇捕手・古巴大賽精神支柱
    "Luis Robert Jr.":      88,   # 天賦全開・競爭心強烈
    # ── 🇵🇷 波多黎各 ────────────────────────────────────────────────────────────
    "Yadier Molina":        97,   # 2013/2017 兩屆 All-WBC・波多黎各精神領袖・大賽魂化身
    "Edwin Díaz":           95,   # 2023 對多明尼加三上三下・稱為生涯 Game 7・賽後受傷仍無悔
    "Carlos Beltrán":       92,   # 2004 季後賽 8 轟 14 打點・2017 WBC .435・名人堂大心臟
    "Carlos Correa":        92,   # 2017 WBC OPS 1.250・世界大賽冠軍・長期大賽球員
    "Francisco Lindor":     92,   # 2017 WBC .370 雙向・全場最活躍球員之一
    "Alex Rios":            89,   # 2013 半決賽 2 分砲打落日本・送波多黎各首次進決賽
    "Roberto Pérez":        87,   # 以捕手守備與心理穩定著稱
    "Martín Maldonado":     86,   # 大賽豐富・心理穩定的後盾型捕手
}

# Position-based speed caps — WBC SB small sample inflates speed for slow-position players
# Individual SPEED_OVERRIDES take priority (use for fast exceptions or extra-slow players)
SPD_POS_CAP: dict[str, int] = {
    "C":  72,   # 捕手通常速度偏慢
    "1B": 70,   # 一壘手通常速度偏慢
    "DH": 65,   # 指定打擊通常速度偏慢
    "3B": 80,   # 三壘手多為重砲手，非跑者型
    "LF": 78,   # 左外野多為強打型，WBC 少量盜壘高估速度
    "RF": 78,   # 右外野同理
}

# Manual speed overrides — based on career SB/yr; corrects WBC small-sample distortion
# Scale: Elite(25+SB)=90-99, Good(15-24)=78-89, Avg(8-14)=68-77, Below(3-7)=62-67, Slow(0-2)=58-61
SPEED_OVERRIDES: dict[str, int] = {
    # ── 極慢（Slow 0–2 SB/yr）──────────────────────────────────────────────────
    "David Ortiz":      50,   # 職涯 17 盜・純 DH・幾乎沒有速度
    "Yadier Molina":    52,   # 全聯盟最慢捕手之一
    "José Abreu":       55,   # 職涯 11 盜・慢速強打一壘手
    "Carlos Delgado":   58,   # 職涯 14 盜・慢速一壘強打
    "Frederich Cepeda": 62,   # 26 年生涯 37 盜・慢但略優於純 DH 型
    "Seung Yuop Lee":   58,   # 純長打一壘手・1–2 盜/季
    # ── 偏慢（Below Average 3–7 SB/yr）─────────────────────────────────────────
    "Albert Pujols":    63,   # 職涯 5 盜/季・偏慢但稍優於純 DH 型
    "Salvador Perez":   60,   # 慢速捕手
    "Yoenis Cespedes":  64,   # 職涯 43 盜・約 5/季・並非速度型
    "Szu-Chi Chou":     64,   # CPBL 約 5 盜/季・偏慢
    "Miguel Cabrera":   65,   # 職涯 40 盜・年輕時中速・晚期極慢・折衷值
    "Freddie Freeman":  62,   # 一壘手・非速度型
    "Paul Goldschmidt": 62,   # 一壘手・中速
    # ── 平均（Average 8–14 SB/yr）──────────────────────────────────────────────
    "Ken Griffey Jr.":  70,   # 職涯 184 盜・約 8/季・平均速度
    "Cheng-Min Peng":   72,   # CPBL 約 11 盜/季・平均速度
    # ── 良好（Good 15–24 SB/yr）────────────────────────────────────────────────
    "Derek Jeter":      82,   # 職涯 358 盜・約 18/季・良好速度
    "Carlos Beltrán":   83,   # 職涯 312 盜・約 16/季・86.4% 成功率
    # ── 慢速位置但實際速度快的例外（覆蓋位置上限）──────────────────────────────
    "Shohei Ohtani":        88,   # DH 但菁英速度・20+ 盜/季
    "J.T. Realmuto":        78,   # MLB 最快捕手・8-12 盜/季
    "Jean Segura":          88,   # 本職 SS・WBC 打 DH・20+ 盜/季
    "Jimmy Rollins":        84,   # 本職 SS・WBC 打 DH・快速跑者
    "Tetsuto Yamada":       82,   # 日本快速二壘/DH・高機動性
    "Carlos Santana":       70,   # C/1B 但速度尚可・約 6 盜/季
    "Yu Chang":             78,   # 1B 但工具型多守位・速度佳
    "Luis Arraez":          75,   # 1B 但快速接觸型打者
    "Bo Gyeong Moon":       80,   # DH 但韓國快速打者
    "Baek Ho Kang":         78,   # DH 但爆發型速度
    # ── 3B/LF/RF 快速例外（覆蓋位置上限）────────────────────────────────────────
    "Manny Machado":        82,   # 3B・運動能力強・生涯約 10 盜/季
    "Yoán Moncada":         80,   # 3B・年輕期快速・15+ 盜/季
    "Christian Yelich":     80,   # LF・巔峰期快速・20 盜/季
    "Kyle Tucker":          76,   # LF・中等速度・約 10 盜/季
    "Jung Hoo Lee":         80,   # RF・快速型外野・KBO 多次盜壘上榜
    # ── 明顯慢速補充（不分位置）────────────────────────────────────────────────
    "Pablo Sandoval":       60,   # 3B・出名慢速・體重問題・生涯 43 盜 / 16 季
    "Juan Soto":            68,   # LF・幾乎不跑・生涯約 3 盜/季
    "Nolan Arenado":        72,   # 3B・平均速度・非跑者型・約 3 盜/季
    "Robinson Canó":        70,   # 2B・生涯約 7 盜/季・非速度型
    "Adam Dunn":            62,   # RF・出名大隻慢・生涯 17 盜・體型龐大
    "Chin-Lung Hu":         83,   # SS・台灣守備型游擊・速度佳但非百米飛人等級
    "Stuart Fairchild":     88,   # CF・有速度但 WBC 小樣本給出 99 失真
    "Tsung-Che Cheng":      88,   # 2B・CPBL 全能球員・速度佳但 WBC 小樣本給出 99 失真
}

# Manual eye overrides — based on career BB% + OBP; corrects WBC small-sample distortion
# Scale: Elite(BB%>12%,OBP>.380)=90-99, Good(9-12%,.360-.380)=80-89, Avg(6-9%,.330-.360)=68-79
EYE_OVERRIDES: dict[str, int] = {
    # ── 頂級選球（Elite）────────────────────────────────────────────────────────
    "Frederich Cepeda": 99,   # 26 年生涯 22% 保送率・.480+ OBP・史上最強選球
    "Cheng-Min Peng":   93,   # CPBL 14% 保送率・.432 OBP・頂級選球（相對 MLB 菁英略調）
    "Albert Pujols":    94,   # 2006 年 14.5% 保送率・.431 OBP・菁英巔峰
    "Carlos Delgado":   92,   # 職涯 12.8% 保送率・.383 OBP・菁英
    "David Ortiz":      92,   # 職涯 13.1% 保送率・.380 OBP・菁英
    "Miguel Cabrera":   88,   # 職涯 10.7% 保送率・.390 OBP・菁英
    # ── 良好選球（Good）──────────────────────────────────────────────────────────
    "Szu-Chi Chou":     84,   # CPBL 約 10% 保送率・.375 OBP・良好（相對 MLB 頂級適當調降）
    "Ken Griffey Jr.":  84,   # 職涯 11.6% 保送率・.370 OBP・良好
    "Derek Jeter":      84,   # 職涯 8.6% 保送率・.377 OBP・2006 年 .417 OBP
    "Carlos Beltrán":   82,   # 職涯 ~11% 保送率・.350 OBP・良好
    "Seung Yuop Lee":   76,   # KBO/NPB 約 8–10% 保送率・良好但非頂級
    # ── 平均選球（Average）───────────────────────────────────────────────────────
    "Omar Vizquel":     72,   # 職涯 8.6% 保送率・.336 OBP・平均
    "Ichiro Suzuki":    76,   # 職涯 6% 保送率但三振率極低・精準接觸・打擊紀律仍屬中上
    # ── 偏低選球（Below Average）─────────────────────────────────────────────────
    "Yoenis Cespedes":  65,   # 職涯 6.9% 保送率・.305 OBP・偏低
    "Yadier Molina":    65,   # 職涯 6.3% 保送率・.327 OBP・偏低
    "Jeremy Peña":      66,   # 職涯 BB%≈4%・著名自由揮棒者・非選球型
    # ── WBC 小樣本壓制（Cap）────────────────────────────────────────────────────
    "Yoel Yanqui":      74,   # CU 2026・無名球員 WBC 小樣本
    "Willy Aybar":      74,   # DO 2009・普通打者・WBC 膨脹（配合 OVR_CEILING）
    "Ernie Clement":    72,   # US 2026・工具人・無選球紀錄
    "Jeff McNeil":      82,   # US 2023・接觸型打者・BB% 中等非 90+ 等級
    "Wander Franco":    84,   # DO 2023・不錯但非頂級選球（配合 MEN 下調）
    "Mike Aviles":      72,   # PR 2017・工具人・WBC 小樣本（配合 OVR_CEILING）
    "Eduardo Perez":    72,   # PR 2006・工具人
    "Taisei Makihara":  74,   # JP 2026・NPB 無名・WBC 小樣本
    "Minho Kang":       72,   # KR 2009・普通捕手・WBC 小樣本
    "Cal Raleigh":      72,   # US 2026・捕手以長打為主非選球型
    "Jose Lopez":       72,   # VE 2009・普通打者（配合 OVR_CEILING）
    "MJ Melendez":      74,   # PR 2026・WBC 小樣本
    "Tetsuto Yamada":   85,   # JP 2023・優秀打者但偏長打型非頂級選球
    "Ha-Seong Kim":     80,   # KR 2017・選球持續進步但 2017 仍在發展中
}

# Manual strikeout/breaking ball overrides (keyed by English player name)
# Use to correct pitchers whose WBC K-rate misrepresents their true profile
K_OVERRIDES: dict[str, int] = {
    # ── 非三振型 — 類型錯誤修正 ──────────────────────────────────────────────
    "Kazuhisa Makita":      62,   # 潛艇式投手・生涯 K/9≈5.5・靠滾地球非三振
    "Hyeon-Jong Yang":      70,   # 接觸型投手・生涯 K/9≈7・誘弱接觸見長
    "Cristopher Sánchez":   65,   # 極端 sinker/GB 投手・生涯 K/9≈6.5・GB% 頂尛
    # ── WBC 小樣本壓制 ───────────────────────────────────────────────────────
    "Masahiro Tanaka":      90,   # 生涯 K/9≈8.5・分叉球犀利但非超頂級三振
    "Chihiro Sumida":       86,   # JP 2026 後援・WBC 小樣本
    "Hiromi Itoh":          86,   # JP 後援・WBC 小樣本
    "Mychal Givens":        82,   # 尚可三振但非頂級・WBC 樣本
    "Michael Wacha":        76,   # 控球/球路投手・非三振型
    "Jose De La Torre":     78,   # PR 2013・WBC 小樣本
    "Pedro Luis Lazo":      80,   # 古巴完投型傳奇・非純三振型
    "Wei-Chung Wang":       80,   # 台灣左投後援・非頂級三振
    "Been Gwak":            78,   # KR 2023・WBC 小樣本
    "Eduardo Rodriguez":    82,   # 生涯 K/9≈9.5・但 WBC 樣本膨脹
    "David Bednar":         92,   # 頂尛終結者 K/9≈10.5・輕微調降
}

# Global OVR shift (set to 0 — distribution now handled via LO/DEFAULT in calc_ratings)
OVR_SHIFT = 0

# Positional stamina baseline (pitchers only)
STAM_BY_POS = {
    "SP": 84, "SP/RP": 78, "RP": 72, "CP": 70, "P": 78,
}

# Minimum OVR floor for world-class players whose WBC sample was too small/unlucky
# Only activates when calc_ratings gives a lower score — never lowers a good score
OVR_FLOOR: dict[str, int] = {
    # ── 🇹🇼 台灣 ──────────────────────────────────────────────────────────────
    "Chien-Ming Wang":      80,   # 洋基雙 19 勝・台灣史上最強投手・小樣本低估
    "Hung-Chih Kuo":        72,   # Dodgers 頂尖後援・2006/2013 IP 過少
    "Szu-Chi Chou":         80,   # 中職安打王・WBC DH 出場但生涯外野守備稱職
    # ── 🇯🇵 日本 ──────────────────────────────────────────────────────────────
    "Ichiro Suzuki":        85,   # 10 座黃金手套・名人堂・任何年份都不應低於此
    "Daisuke Matsuzaka":    80,   # 2006/2009 WBC MVP・頂尖先發
    "Kosuke Fukudome":      75,   # NPB 多座黃金手套・MLB 固定先發
    "Norichika Aoki":       72,   # NPB 多次打擊王・MLB 固定先發・樣本過少
    "Hayato Sakamoto":      78,   # NPB 10 座黃金手套游擊・2013 WBC 打擊失常
    "Kyuji Fujikawa":       75,   # NPB 史上最強終結者之一・樣本過少
    "Masahiro Tanaka":      78,   # 24-0 賽季・Game 7 英雄・WBC 表現被低估
    "Hideaki Wakui":        72,   # NPB 多次 10 勝王牌・1.2 IP 失真
    "Roki Sasaki":          80,   # NPB 完全比賽投手・世界頂尖潛力球員
    "Tomoyuki Sugano":      76,   # 5 次 NPB 防禦率王・NPB 頂尖先發
    "Yoshinobu Yamamoto":   84,   # Dodgers 王牌・$3.25 億合約・2026 WBC 僅 6.2 IP
    "Munetaka Murakami":    74,   # NPB 三冠王・單季 56 轟紀錄・WBC 偶爾失常
    "Kensuke Kondoh":       72,   # NPB 頂級打者・2026 WBC 0 安失真
    "Shugo Maki":           70,   # NPB 頂級二壘手・WBC 樣本小
    "Kazuma Okamoto":       70,   # NPB 年輕強打・2023 WBC 已交出 92 分
    "Yusei Kikuchi":        74,   # MLB 先發・WBC ERA 失真
    "Kaito Kozono":         68,   # NPB 新生代游擊明星・4 PA 不代表實力
    "Ryosuke Kikuchi":      72,   # 10 連座 NPB 黃金手套・2017 WBC 打擊不佳
    "Seiya Suzuki":         70,   # 2017 WBC 表現差但已是 NPB 頂級外野
    "Yu Darvish":           82,   # 多次 Cy Young 票選・未來名人堂・2023 WBC 6 IP 失真
    # ── 🇰🇷 韓國 ──────────────────────────────────────────────────────────────
    "Hyun Jin Ryu":         76,   # KBO/MLB 頂尖左投・多年 All-Star・WBC ERA 失真
    "Seunghwan Oh":         76,   # 石佛・KBO 400+ 救援・1 IP 樣本完全失真
    "Byung-Hyun Kim":       72,   # MLB 知名後援/先發・WBC 樣本小
    "Chang-Yong Lim":       74,   # KBO 史上最強終結者之一・WBC 5.1 IP 失真
    "Ha-Seong Kim":         76,   # MLB + KBO 黃金手套・2017 WBC 5 PA 失真
    "Jung Hoo Lee":         78,   # KBO 打擊王・Giants 大合約・2026 WBC 失常
    "Hyeseong Kim":         72,   # KBO 4 座黃金手套・2026 WBC 樣本失真
    "Sung Bum Na":          72,   # KBO 全明星強打・6 PA 失真
    # ── 🇺🇸 美國 ──────────────────────────────────────────────────────────────
    "Chipper Jones Jr.":    80,   # 名人堂・史上最強開關打者之一・11 PA 0 安完全失真
    "Matt Holliday":        76,   # 多次 All-Star・6 PA 0 安失真
    "Dustin Pedroia":       78,   # AL MVP・4 座黃金手套・WBC 失常
    "Paul Goldschmidt":     82,   # NL MVP・多次 All-Star・16 PA .077 失真
    "Nolan Arenado":        82,   # 10 連座黃金手套・33 PA 時期仍是頂尖
    "Aaron Judge":          82,   # AL HR 紀錄保持人・MVP・WBC 失常
    "Bryce Harper":         82,   # 兩次 MVP・費城人王牌・30 PA 失真
    "Alex Bregman":         78,   # 兩次 World Series 冠軍・頂尖三壘手・WBC 失常
    "Daniel Murphy":        72,   # NLCS 英雄・全明星水準・6 PA 失真
    "Bobby Witt Jr.":       78,   # MLB 頂尖新生代 SS・3 PA 失真
    "Pete Alonso":          76,   # 兩次 HR Derby 冠軍・15 PA 失真
    "Craig Kimbrel":        76,   # 顛峰期最強終結者之一・3.2 IP 失真
    "Andrew Miller":        80,   # 2016 ALCS MVP・最強後援之一・K/9=12.7 史上頂尛・2.2 IP 失真
    "Byron Buxton":         76,   # 5 工具明星 CF・9 PA 0 安失真
    "Mark Teixeira":        78,   # 5 座黃金手套一壘手・15 PA 0 安失真
    # ── 🇩🇴 多明尼加 ────────────────────────────────────────────────────────
    "Alfonso Soriano":      76,   # MLB 五工具球星・多次 All-Star・15 PA 0 安失真
    "Robinson Canó":        80,   # 多次 All-Star・黃金手套・2009 WBC 14 PA 失真
    "José Reyes":           78,   # 打擊王・多次 All-Star・各屆 WBC 樣本失真
    "Manny Machado":        82,   # 多次黃金手套・大合約・WBC 表現不穩
    "Sandy Alcantara":      78,   # NL Cy Young 得主・WBC 3-4 IP 失真
    "Rafael Devers":        78,   # 多次 All-Star・大合約・18 PA 失真
    "Gary Sánchez":         72,   # MLB 強打捕手・多次 All-Star・6 PA 失真
    "Teoscar Hernández":    75,   # Silver Slugger・7 PA 失真
    "Julio Rodríguez":      78,   # 西雅圖明星・新生代頂尖球員・WBC 失常
    "Hanley Ramirez":       76,   # NL 打擊王・MLB 頂尖 SS 之一・12 PA 失真
    "Adrian Beltré":        84,   # 名人堂・史上最強守備三壘手・任何年份不應低於此
    # ── 🇻🇪 委內瑞拉 ────────────────────────────────────────────────────────
    "Miguel Cabrera":       84,   # 三冠王・名人堂・史上最強右打之一・任何年份不應低於此
    "Magglio Ordonez":      76,   # MLB 打擊王・多次 All-Star・21 PA 失真
    "Jose Altuve":          82,   # AL MVP・多次 All-Star・任何年份不應低於此
    "Ronald Acuña Jr.":     84,   # NL MVP・五工具頂尖球員・WBC 完全失常
    "Félix Hernández":      80,   # Cy Young・生涯頂尖先發・WBC 成績不代表能力
    "Bobby Abreu":          76,   # 多次 All-Star・OBP 頂尖・26 PA 低打率失真
    "Salvador Perez":       76,   # 6 座黃金手套捕手・WBC 打擊失常
    "Gleyber Torres":       72,   # MLB All-Star・13 PA 失真
    "Francisco Rodríguez":  74,   # 單季 62 救援紀錄・WBC 2.1 IP 失真
    "Aníbal Sánchez":       74,   # NL 防禦率王・0.1 IP 完全失真
    "Carlos Zambrano":      76,   # Cubs 王牌・多次 All-Star・WBC ERA 失真
    "Andrés Giménez":       72,   # 3 連座黃金手套・WBC 樣本失真
    # ── 🇨🇺 古巴 ────────────────────────────────────────────────────────────
    "Aroldis Chapman":      78,   # MLB 最速球投手・WBC 古巴時期展現超強潛力
    "Yoenis Cespedes":      75,   # MLB All-Star・8 PA 失真
    "Alfredo Despaigne":    72,   # 古巴/NPB 頂尖強打・WBC 失常
    "Raisel Iglesias":      75,   # 古巴/MLB 頂尖後援・4.2 IP 失真
    "Luis Robert Jr.":      78,   # MLB 黃金手套 CF・28 PA .259 但 OVR 嚴重低估
    # ── 🇵🇷 波多黎各 ────────────────────────────────────────────────────────
    "Yadier Molina":        82,   # 名人堂捕手・9 座黃金手套・任何年份不應低於此
    "Ivan Rodriguez":       82,   # 名人堂・13 座黃金手套・史上最強守備捕手
    "Carlos Beltrán":       80,   # 準名人堂・5 工具・WBC 28-37 PA 失真
    "Carlos Delgado":       80,   # 470+ HR・PR 最強一壘手・1 PA 完全失真
    "Edwin Díaz":           80,   # 2022 最強終結者・WBC 2 IP 樣本失真
    "Alexis Díaz":          74,   # MLB 終結者・2 IP 失真
    "José Berríos":         74,   # MLB 頂尖先發・2 IP 失真・WBC 小樣本
    "Seth Lugo":            75,   # MLB 先發・4.1 IP ERA 失真
    "Carlos Correa":        80,   # World Series 冠軍・WBC OVR 有時失真
    "Francisco Lindor":     80,   # 多次黃金手套・WBC OVR 有時失真
}

# OVR ceiling for players whose WBC small sample inflated their rating above realistic level
# Only activates when calc_ratings gives a higher score — never raises a low score
OVR_CEILING: dict[str, int] = {
    # ── 🇻🇪 委內瑞拉 ────────────────────────────────────────────────────────────
    "Henry Blanco":         78,   # 後備捕手・WBC 小樣本打出高 OPS 失真
    "Jose Lopez":           76,   # 平均水準 MLB 二壘手・WBC 樣本嚴重高估
    # ── 🇩🇴 多明尼加 ────────────────────────────────────────────────────────────
    "Willy Aybar":          74,   # 工具型內野手・非主力球員
    "Gregory Polanco":      84,   # 好但非頂尖・WBC 少量 AB 爆炸失真
    # ── 🇵🇷 波多黎各 ────────────────────────────────────────────────────────────
    "Mike Aviles":          76,   # 工具型替補球員・非明星水準
    # ── 🇺🇸 美國 ──────────────────────────────────────────────────────────────
    "Chris Iannetta":       78,   # 中等捕手・WBC 少量 AB 高打率失真
    "Brian Roberts":        82,   # 藥檢期間・WBC 樣本高估
    # ── 🇰🇷 韓國 ──────────────────────────────────────────────────────────────
    "Sung Heon Hong":       80,   # 非明星捕手・WBC 小樣本高估
}

POS_ZH = {
    "C": "捕手", "1B": "一壘手", "2B": "二壘手", "3B": "三壘手", "SS": "游擊手",
    "LF": "左外野手", "CF": "中外野手", "RF": "右外野手", "OF": "外野手", "DH": "指定打擊",
    "SP": "先發投手", "RP": "後援投手", "CP": "終結者", "SP/RP": "先發兼後援",
    "P": "投手", "INF": "內野手", "UT": "工具人",
}

SECTION_ORDER = [
    "Catchers", "Infielders", "Outfielders",
    "Designated Hitter", "Two-Way Players", "Pitchers",
]
SECTION_ZH = {
    "Catchers":          "捕手 (C)",
    "Infielders":        "內野手",
    "Outfielders":       "外野手",
    "Designated Hitter": "指定打擊 (DH)",
    "Two-Way Players":   "二刀流",
    "Pitchers":          "投手",
}

# ── 傳奇球員一般版（rar 非 x，給沒有 WBC 出賽記錄的傳奇球員）────────────────
# 這些球員只有傳奇卡，補上一般版讓玩家可以取得非傳奇版本
# stats 野手:[打擊力,選球眼,速度,守備力,心理]
REGULAR_EXTRAS = [
]

# ── 傳奇球員（rar:'x'，手動維護）────────────────────────────────────────────
# stats 野手:[打擊力,選球眼,速度,守備力,心理]  投手:[投球力,控球力,變化球,體力,心理]
LEGENDS = [
    # 🇹🇼 台灣
    "{name:'彭政閔 [2013 傳奇]',year:2013,en:'Cheng-Min Peng',nat:'🇹🇼',pos:'1B',av:'🐘',rar:'x',ovr:99,pit:false,era:[2013],stats:[99,97,78,95,99],story:'<hl>復古傳奇特別款</hl>・彭政閔・中職先生・2013 WBC 以肉身擋球的鐵血精神感動全台・永遠的中華隊隊長。',skills:[{i:'👑',n:'鐵血隊長',d:'守備+20%'},{i:'🛡️',n:'護國神柱',d:'打擊無上限'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'王建民 [2013 傳奇]',year:2013,en:'Chien-Ming Wang',nat:'🇹🇼',pos:'SP',av:'⬇️',rar:'x',ovr:99,pit:true,era:[2013],stats:[99,99,99,90,99],story:'<hl>復古傳奇特別款</hl>・王建民・2013 WBC 完美鎮壓・那個年代全台灣最熟悉的名字・每一顆伸卡球都是台灣棒球不滅的傳說。',skills:[{i:'⬇️',n:'伸卡地獄',d:'地滾球率永恆第一'},{i:'📸',n:'復古印記',d:'歷史加成+30%'},{i:'🇹🇼',n:'台灣驕傲',d:'全開無限制'}]}",
    "{name:'周思齊 [2013 傳奇]',year:2013,en:'Szu-Chi Chou',nat:'🇹🇼',pos:'LF',av:'🎯',rar:'x',ovr:99,pit:false,era:[2013],stats:[99,99,72,78,99],story:'<hl>復古傳奇特別款</hl>・周思齊・2013 WBC 對決田中將大敲出超前安打・那一刻全台灣屏息・中職最強安打機器。',skills:[{i:'🎯',n:'安打之神',d:'打擊率永恆頂點'},{i:'📸',n:'復古印記',d:'歷史加成+30%'},{i:'🔥',n:'大心臟',d:'落後時打擊無上限'}]}",
    # 🇯🇵 日本
    "{name:'鈴木一朗 [2006 傳奇]',year:2006,en:'Ichiro Suzuki',nat:'🇯🇵',pos:'CF',av:'🎖️',rar:'x',ovr:99,pit:false,era:[2006],stats:[99,88,99,99,99],story:'<hl>復古傳奇特別款</hl>・鈴木一朗・安打製造機・2006 WBC 帶領日本奪冠・棒球史上最純粹的打擊藝術家。',skills:[{i:'🎖️',n:'安打紀錄',d:'安打數永恆第一'},{i:'💨',n:'超音速跑壘',d:'速度無上限'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'松坂大輔 [2006 傳奇]',year:2006,en:'Daisuke Matsuzaka',nat:'🇯🇵',pos:'SP',av:'🌀',rar:'x',ovr:99,pit:true,era:[2006],stats:[99,96,99,88,99],story:'<hl>復古傳奇特別款</hl>・松坂大輔・2006 WBC 最強先發・螺旋球讓全球打者束手無策・世代級王牌的絕對壓制。',skills:[{i:'🌀',n:'螺旋球',d:'揮空率無上限'},{i:'🔥',n:'WBC MVP',d:'關鍵局壓制+50%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'大谷翔平 [2023 傳奇]',year:2023,en:'Shohei Ohtani',nat:'🇯🇵',pos:'SP',av:'🌟',rar:'x',ovr:99,pit:true,era:[2023],stats:[99,99,99,95,99],story:'<hl>復古傳奇特別款</hl>・大谷翔平・2023 WBC MVP・決賽最後一球親手三振 Mike Trout・賽前演講「今天讓我們憧憬大聯盟吧」・史上最完整的二刀流傳奇。',skills:[{i:'🌟',n:'二刀流',d:'投打全數據無上限'},{i:'🔥',n:'決賽終結者',d:'關鍵局壓制+99%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    # 🇰🇷 韓國
    "{name:'李承燁 [2006 傳奇]',year:2006,en:'Seung-yeop Lee',nat:'🇰🇷',pos:'1B',av:'💣',rar:'x',ovr:99,pit:false,era:[2006],stats:[82,99,75,78,99],story:'<hl>復古傳奇特別款</hl>・李承燁・韓國全壘打王・KBO 史上最恐怖的長打威脅・2006 WBC 核心砲手。',skills:[{i:'💣',n:'全壘打王',d:'HR無上限'},{i:'👑',n:'韓國英雄',d:'全隊士氣+20%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'朴贊浩 [2006 傳奇]',year:2006,en:'Chan Ho Park',nat:'🇰🇷',pos:'SP',av:'🔥',rar:'x',ovr:99,pit:true,era:[2006],stats:[99,92,99,88,99],story:'<hl>復古傳奇特別款</hl>・朴贊浩・韓國旅美先驅・2006 WBC 以壓倒性球威代表韓國出戰・亞洲投手進軍 MLB 的燈塔。',skills:[{i:'🔥',n:'旅美先驅',d:'球速無上限'},{i:'⚡',n:'壓制力',d:'三振率+40%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'柳賢振 [2013 傳奇]',year:2013,en:'Hyun-Jin Ryu',nat:'🇰🇷',pos:'SP',av:'🎯',rar:'x',ovr:99,pit:true,era:[2013],stats:[96,99,92,88,99],story:'<hl>復古傳奇特別款</hl>・柳賢振・2013 WBC 韓國王牌・精密控球配合多樣變化球，壓制各國強打。',skills:[{i:'🎯',n:'精準控球',d:'四壞球率永久降低'},{i:'🌀',n:'變化球大師',d:'揮空率+40%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    # 🇺🇸 美國
    "{name:'Derek Jeter [2006 傳奇]',year:2006,en:'Derek Jeter',nat:'🇺🇸',pos:'SS',av:'👑',rar:'x',ovr:99,pit:false,era:[2006],stats:[90,85,88,92,99],story:'<hl>復古傳奇特別款</hl>・Derek Jeter・洋基隊長・2006 WBC 美國隊精神領袖・名人堂等級的游擊手。',skills:[{i:'👑',n:'洋基隊長',d:'全隊 OVR+15%'},{i:'🛡️',n:'黃金手套',d:'守備無上限'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'Ken Griffey Jr. [2006 傳奇]',year:2006,en:'Ken Griffey Jr.',nat:'🇺🇸',pos:'CF',av:'🦅',rar:'x',ovr:99,pit:false,era:[2006],stats:[88,99,88,95,99],story:'<hl>復古傳奇特別款</hl>・Ken Griffey Jr.・The Kid・完美揮棒弧度・2006 WBC 史上最全能的中外野手之一。',skills:[{i:'🦅',n:'完美揮棒',d:'HR+無上限'},{i:'💨',n:'中外野霸主',d:'守備+速度無上限'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'Marcus Stroman [2017 傳奇]',year:2017,en:'Marcus Stroman',nat:'🇺🇸',pos:'SP',av:'✊',rar:'x',ovr:99,pit:true,era:[2017],stats:[88,92,85,88,99],story:'<hl>復古傳奇特別款</hl>・Marcus Stroman・2017 WBC 美國奪冠功臣・小個子大心臟・關鍵時刻永不退縮的鬥士。',skills:[{i:'✊',n:'鬥士精神',d:'關鍵局壓制無上限'},{i:'⬇️',n:'地滾球製造機',d:'雙殺機率+40%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    # 🇩🇴 多明尼加
    "{name:'Albert Pujols [2006 傳奇]',year:2006,en:'Albert Pujols',nat:'🇩🇴',pos:'1B',av:'🔱',rar:'x',ovr:99,pit:false,era:[2006],stats:[88,99,88,78,99],story:'<hl>復古傳奇特別款</hl>・Albert Pujols・機器人・2006 WBC 多明尼加核心・史上最完整的打者之一。',skills:[{i:'🔱',n:'機器人',d:'得點圈打點無上限'},{i:'💣',n:'長打之王',d:'HR無上限'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'David Ortiz [2006 傳奇]',year:2006,en:'David Ortiz',nat:'🇩🇴',pos:'DH',av:'🦁',rar:'x',ovr:99,pit:false,era:[2006],stats:[85,99,72,62,99],story:'<hl>復古傳奇特別款</hl>・David Ortiz・Big Papi・2006 WBC 多明尼加代表・最令投手恐懼的指定打擊。',skills:[{i:'🦁',n:'Big Papi',d:'得點圈打擊無上限'},{i:'🔥',n:'逆境英雄',d:'落後時 OVR+30%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'Pedro Martínez [2009 傳奇]',year:2009,en:'Pedro Martínez',nat:'🇩🇴',pos:'SP',av:'🌟',rar:'x',ovr:99,pit:true,era:[2009],stats:[99,96,99,82,99],story:'<hl>復古傳奇特別款</hl>・Pedro Martínez・2009 WBC 多明尼加傳奇王牌・三冠王級的數據與壓倒性的支配力。',skills:[{i:'🌟',n:'支配者',d:'OPS 壓制永恆頂點'},{i:'💨',n:'三振藝術',d:'三振率無上限'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    # 🇻🇪 委內瑞拉
    "{name:'Miguel Cabrera [2006 傳奇]',year:2006,en:'Miguel Cabrera',nat:'🇻🇪',pos:'3B',av:'🏆',rar:'x',ovr:99,pit:false,era:[2006],stats:[92,99,85,72,99],story:'<hl>復古傳奇特別款</hl>・Miguel Cabrera・Miggy・三冠王・2006 WBC 委內瑞拉最強打者・史上最偉大的右打之一。',skills:[{i:'🏆',n:'三冠王',d:'打擊全數據無上限'},{i:'💣',n:'右打之神',d:'HR無上限'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'Omar Vizquel [2006 傳奇]',year:2006,en:'Omar Vizquel',nat:'🇻🇪',pos:'SS',av:'✨',rar:'x',ovr:99,pit:false,era:[2006],stats:[82,72,85,99,97],story:'<hl>復古傳奇特別款</hl>・Omar Vizquel・11 座 MLB 黃金手套・史上守備最優雅的游擊手・委內瑞拉棒球永恆象徵・每一次接球都是藝術。',skills:[{i:'✨',n:'守備藝術',d:'守備無上限'},{i:'💎',n:'黃金手套之神',d:'失誤率永久清零'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'Félix Hernández [2009 傳奇]',year:2009,en:'Félix Hernández',nat:'🇻🇪',pos:'SP',av:'👑',rar:'x',ovr:99,pit:true,era:[2009],stats:[99,94,97,88,99],story:'<hl>復古傳奇特別款</hl>・Félix Hernández・委內瑞拉之王・Cy Young 得主・2009 WBC 親口說是生涯最重要一場・8.2 局封鎖幫委內瑞拉晉級・以壓倒性球威主宰打線。',skills:[{i:'👑',n:'委內瑞拉之王',d:'投球力無上限'},{i:'🔥',n:'Cy Young 等級',d:'防禦率永久最低'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    # 🇨🇺 古巴
    "{name:'Frederich Cepeda [2006 傳奇]',year:2006,en:'Frederich Cepeda',nat:'🇨🇺',pos:'LF',av:'🦅',rar:'x',ovr:99,pit:false,era:[2006],stats:[95,90,82,82,99],story:'<hl>復古傳奇特別款</hl>・Frederich Cepeda・古巴棒球象徵・三屆 WBC 常勝將軍・整個世代最恐怖的古巴打者。',skills:[{i:'🦅',n:'古巴之鷹',d:'打擊率永恆頂點'},{i:'🔥',n:'三屆老將',d:'關鍵打擊+40%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'Yoenis Cespedes [2009 傳奇]',year:2009,en:'Yoenis Cespedes',nat:'🇨🇺',pos:'CF',av:'💨',rar:'x',ovr:99,pit:false,era:[2009],stats:[88,95,82,95,99],story:'<hl>復古傳奇特別款</hl>・Yoenis Cespedes・2013 WBC 古巴砲手・驚人的工具型天賦・速度與長打並存的全能外野手。',skills:[{i:'💨',n:'工具型天才',d:'速度+長打無上限'},{i:'💣',n:'砲轟外野',d:'HR+40%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'José Abreu [2013 傳奇]',year:2013,en:'José Abreu',nat:'🇨🇺',pos:'1B',av:'💪',rar:'x',ovr:99,pit:false,era:[2013],stats:[85,99,72,78,99],story:'<hl>復古傳奇特別款</hl>・José Abreu・2013 WBC 古巴核心一壘手・赴美後立刻成為 MLB 頂尖強打・古巴棒球的最後傳承。',skills:[{i:'💪',n:'古巴傳承',d:'長打率無上限'},{i:'🔱',n:'新人王',d:'OVR永恆99'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    # 🇵🇷 波多黎各
    "{name:'Carlos Beltrán [2006 傳奇]',year:2006,en:'Carlos Beltrán',nat:'🇵🇷',pos:'CF',av:'⭐',rar:'x',ovr:99,pit:false,era:[2006],stats:[90,88,90,92,99],story:'<hl>復古傳奇特別款</hl>・Carlos Beltrán・波多黎各 WBC 精神隊長・攻守跑三項全能・中外野史上最完整的球員之一。',skills:[{i:'⭐',n:'五工具球員',d:'全數據無上限'},{i:'👑',n:'波多黎各隊長',d:'全隊 OVR+15%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'Carlos Delgado [2009 傳奇]',year:2009,en:'Carlos Delgado',nat:'🇵🇷',pos:'1B',av:'🔱',rar:'x',ovr:99,pit:false,era:[2009],stats:[88,99,88,72,99],story:'<hl>復古傳奇特別款</hl>・Carlos Delgado・2009 WBC 波多黎各最強打者・左打長砲・職涯超過 470 支全壘打的恐怖打線核心。',skills:[{i:'🔱',n:'左打大砲',d:'HR無上限'},{i:'💣',n:'波多黎各砲手',d:'得點圈打點+40%'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
    "{name:'Yadier Molina [2006 傳奇]',year:2006,en:'Yadier Molina',nat:'🇵🇷',pos:'C',av:'🛡️',rar:'x',ovr:99,pit:false,era:[2006],stats:[85,80,78,99,99],story:'<hl>復古傳奇特別款</hl>・Yadier Molina・史上最佳捕手・2006 WBC 波多黎各守護者・阻殺率與引導能力永恆登峰。',skills:[{i:'🛡️',n:'本壘守護神',d:'阻殺率永恆頂點'},{i:'🧠',n:'捕手之王',d:'投手 ERA 永久-1'},{i:'📸',n:'復古印記',d:'歷史加成+30%'}]}",
]


# ── helpers ──────────────────────────────────────────────────────────────────

def safe_int(v, default: int = 65) -> int:
    try:
        f = float(v)
        return default if math.isnan(f) else int(round(f))
    except (TypeError, ValueError):
        return default


def clamp(v, lo: int = 40, hi: int = 99) -> int:
    return max(lo, min(hi, int(round(v))))


def ovr_to_rar(ovr: int) -> str:
    if ovr >= 90: return "h"   # 頂尖  ~9%
    if ovr >= 80: return "l"   # 精英 ~30%
    if ovr >= 69: return "r"   # 強力 ~31%
    return "c"                  # 一般 ~30%


def js_str(s: str) -> str:
    s = str(s).replace("\\", "\\\\").replace("'", "\\'")
    return f"'{s}'"


BAT_EXCEL_FIELDS = [
    ("BAT_G", "g"), ("BAT_AB", "ab"), ("BAT_R", "r"), ("BAT_H", "h"),
    ("BAT_2B", "d2"), ("BAT_3B", "d3"), ("BAT_HR", "hr"), ("BAT_RBI", "rbi"),
    ("BAT_BB", "bb"), ("BAT_SO", "so"), ("BAT_SB", "sb"), ("BAT_CS", "cs"),
    ("BAT_AVG", "avg"), ("BAT_OBP", "obp"), ("BAT_SLG", "slg"), ("BAT_OPS", "ops"),
    ("BAT_PA", "pa"), ("BAT_HBP", "hbp"), ("BAT_SAC", "sac"), ("BAT_SF", "sf"),
    ("BAT_GIDP", "gidp"), ("BAT_GO/AO", "goao"), ("BAT_XBH", "xbh"), ("BAT_TB", "tb"),
    ("BAT_IBB", "ibb"), ("BAT_BABIP", "babip"), ("BAT_ISO", "iso"), ("BAT_AB/HR", "abhr"),
    ("BAT_BB/K", "bbk"), ("BAT_BB%", "bbpct"), ("BAT_K%", "kpct"),
]

PIT_EXCEL_FIELDS = [
    ("PIT_W", "w"), ("PIT_L", "l"), ("PIT_ERA", "era"), ("PIT_G", "g"),
    ("PIT_GS", "gs"), ("PIT_CG", "cg"), ("PIT_SHO", "sho"), ("PIT_SV", "sv"),
    ("PIT_SVO", "svo"), ("PIT_IP", "ip"), ("PIT_H", "h"), ("PIT_R", "r"),
    ("PIT_ER", "er"), ("PIT_HR", "hr"), ("PIT_HB", "hb"), ("PIT_BB", "bb"),
    ("PIT_SO", "so"), ("PIT_WHIP", "whip"), ("PIT_AVG", "avg"), ("PIT_TBF", "tbf"),
    ("PIT_NP", "np"), ("PIT_P/IP", "pip"), ("PIT_QS", "qs"), ("PIT_GF", "gf"),
    ("PIT_HLD", "hld"), ("PIT_IBB", "ibb"), ("PIT_WP", "wp"), ("PIT_BK", "bk"),
    ("PIT_GDP", "gdp"), ("PIT_GO/AO", "goao"), ("PIT_SO/9", "so9"), ("PIT_BB/9", "bb9"),
    ("PIT_K/BB", "kbb"), ("PIT_BABIP", "babip"), ("PIT_SB", "sb"), ("PIT_CS", "cs"),
    ("PIT_PK", "pk"),
]


def js_val(v) -> str:
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    return js_str(v)


def js_obj(d: dict) -> str:
    parts = [f"{k}:{js_val(v)}" for k, v in d.items()]
    return "{" + ",".join(parts) + "}"


def excel_stats(d: dict, is_pitcher: bool) -> dict:
    pairs = PIT_EXCEL_FIELDS if is_pitcher else BAT_EXCEL_FIELDS
    out = {}
    for src, alias in pairs:
      val = d.get(src)
      if val is None or val == "":
          continue
      out[alias] = val
    return out


# ── stat arrays ───────────────────────────────────────────────────────────────

def bat_stats(d: dict) -> list[int]:
    """[打擊力, 選球眼, 速度, 守備力, 心理]"""
    pos = d["pos"] or "OF"
    bat = safe_int(d["RTG_BAT"])
    en  = d.get("name_en") or ""
    eye = EYE_OVERRIDES[en]   if en in EYE_OVERRIDES   else safe_int(d["RTG_EYE"])
    if en in SPEED_OVERRIDES:
        spd = SPEED_OVERRIDES[en]
    else:
        spd = safe_int(d["RTG_SPEED"])
        cap = SPD_POS_CAP.get(pos, 99)
        spd = min(spd, cap)
    ovr = safe_int(d["RTG_OVR"])
    # Use manual override for famous defenders; else derive from position + OVR
    if en in DEF_OVERRIDES:
        dfs = DEF_OVERRIDES[en]
    else:
        dfs = clamp(DEF_BY_POS.get(pos, 72) + (ovr - 70) // 6)
    yr  = d.get("season") or ""
    men_key = (en, int(yr)) if (en, int(yr) if yr else 0) in MEN_OVERRIDES else en
    if men_key in MEN_OVERRIDES:
        men = MEN_OVERRIDES[men_key]
    else:
        men = clamp(58 + (ovr - 40) * 42 // 59)   # 40→58, 99→99
    return [bat, eye, spd, dfs, men]


def bat_power(d: dict) -> int:
    """力量（獨立欄位，來源：RTG_POWER）"""
    return safe_int(d["RTG_POWER"])


def pit_stats(d: dict) -> list[int]:
    """[投球力, 控球力, 變化球, 體力, 心理]"""
    pos  = d["pos"] or "RP"
    pit  = safe_int(d["RTG_PIT"])
    ctrl = safe_int(d["RTG_CTRL"])
    en   = d.get("name_en") or ""
    k    = K_OVERRIDES[en] if en in K_OVERRIDES else safe_int(d["RTG_K"])
    ovr  = safe_int(d["RTG_OVR"])
    en   = d.get("name_en") or ""
    stam = clamp(STAM_BY_POS.get(pos, 72) + (ovr - 70) // 6)
    yr  = d.get("season") or ""
    men_key = (en, int(yr)) if (en, int(yr) if yr else 0) in MEN_OVERRIDES else en
    if men_key in MEN_OVERRIDES:
        men = MEN_OVERRIDES[men_key]
    else:
        men = clamp(58 + (ovr - 40) * 42 // 59)
    return [pit, ctrl, k, stam, men]


# ── skills ────────────────────────────────────────────────────────────────────

def make_skills(d: dict, is_pitcher: bool) -> list[dict]:
    nat  = d["nat"]
    flag = NAT_TO_FLAG.get(nat, "🌐")
    ovr  = safe_int(d["RTG_OVR"])
    skills: list[dict] = []

    if is_pitcher:
        ctrl = safe_int(d["RTG_CTRL"])
        k    = safe_int(d["RTG_K"])
        era  = safe_int(d["RTG_ERA_Q"])
        pos  = d["pos"] or "RP"

        # Skill 1: dominant quality
        if k >= ctrl and k > 70:
            skills.append({"i": "💨", "n": "三振利器",    "d": f"三振率+{max(5,(k-65)//4)}%"})
        elif ctrl > 70:
            skills.append({"i": "🎯", "n": "精密控球",    "d": "保送率降低"})
        elif era > 70:
            skills.append({"i": "🧊", "n": "壓制型投球",  "d": "失分抑制+8%"})
        else:
            skills.append({"i": "⚾", "n": "穩定出賽",    "d": "體力消耗-8%"})

        # Skill 2: role
        if pos in ("SP", "SP/RP"):
            bonus = max(8, (ovr - 70) // 3)
            skills.append({"i": "🔥", "n": "先發王牌",  "d": f"投球品質+{bonus}%"})
        elif pos == "CP":
            skills.append({"i": "🔒", "n": "終結者",    "d": "救援成功率+12%"})
        else:
            skills.append({"i": "⚡", "n": "中繼橋樑",  "d": "連投體力+8%"})

    else:
        pwr = safe_int(d["RTG_POWER"])
        spd = safe_int(d["RTG_SPEED"])
        eye = safe_int(d["RTG_EYE"])
        con = safe_int(d["RTG_CONTACT"])

        candidates = [
            (pwr, "💣", "長打砲火",   f"HR+{max(5,(pwr-65)//4)}%"),
            (spd, "⚡", "速度型球員", f"盜壘成功率+{max(5,(spd-65)//4)}%"),
            (eye, "👁️", "選球高手",   f"上壘率+{max(5,(eye-65)//4)}%"),
            (con, "🎯", "精準打擊",   f"打擊率+{max(5,(con-65)//4)}%"),
        ]
        candidates.sort(reverse=True, key=lambda x: x[0])
        for _, icon, name, desc in candidates[:2]:
            skills.append({"i": icon, "n": name, "d": desc})

    skills.append({"i": flag, "n": nat, "d": "+5%"})
    return skills


# ── story overrides ───────────────────────────────────────────────────────────
# Keys: (name_en, year:int) for year-specific, or name_en for cross-year fallback
# Stories in Traditional Chinese, 1–3 sentences, factual only.

STORY_OVERRIDES: dict = {
    # ══════════════════════════════════════════════════════════════════
    # 🇹🇼 台灣
    # ══════════════════════════════════════════════════════════════════

    # ── 🇹🇼 2006 ──────────────────────────────────────────────────────
    ("Fong-Ming Chen",    2006): "2006 WBC 台灣代表隊主力捕手，中職生涯以穩健配球著稱，與葉君璋並列為台灣最具代表性的捕手之一。",
    ("Chun-Chang Yeh",    2006): "中職史上唯一出賽破千場的捕手，2006 WBC 是其多次代表台灣出征國際賽的縮影，生涯橫跨雅典、北京兩屆奧運。",
    ("Yung-Chi Chen",     2006): "2006 WBC 台灣代表隊三壘手，返台後縱橫中職 16 個球季並累積 136 支全壘打，成為統一獅隊史全壘打王。",
    ("Chin-Lung Hu",      2006): "2006 WBC 台灣代表隊游擊手，翌年成為台灣首位在 MLB 擊出全壘打的野手，入選道奇首發後第二次打席即送出全壘打，曾被視為台灣旅美游擊最大希望。",
    ("Chih-Sheng Lin",    2006): "2006 WBC 台灣代表隊一壘手，中職生涯 2023 年成為史上首位累積 300 支全壘打的球員，是台灣職棒全壘打紀錄的新保持人。",
    ("Chang-Ming Cheng",  2006): "2006 WBC 台灣代表隊內野手，以守備穩健著稱的中職老將，是第一屆世界棒球經典賽台灣陣容中的可靠內野支柱。",
    ("Chia-Hao Chang",    2006): "2006 WBC 台灣代表隊一壘手，中職生涯三壘安打數曾為台灣職棒史上最多，退役後轉任守備與跑壘教練。",
    ("Dai-Kang Yang",     2006): "2006 WBC 台灣代表隊球員，同年以高中生身份成為日職選秀狀元加入北海道日本火腿鬥士，旅日生涯出賽逾 1300 場並轟出 105 支全壘打，是台日兩地球迷心目中的國民英雄。",
    ("Shen Yang",         2006): "2005 年以 121 支安打同時拿下中職打擊王與盜壘王，以出色的接觸打擊與速度入選第一屆世界棒球經典賽。",
    ("Tai-Shan Chang",    2006): "中職 20 年生涯轟出 289 支全壘打，入選 2006 WBC 時已是台灣職棒全壘打紀錄保持人，是本屆陣容中最具長打威脅的指定打擊。",
    ("Chien-Ming Chang",  2006): "2006 WBC 台灣代表隊右外野手，以「雷射砲手」聞名，臂力超強並累積中職外野史上最多助殺紀錄之一，多次榮獲金手套獎。",
    ("Chia-Hsien Hseih",  2006): "2006 WBC 台灣代表隊中外野手，被譽為台灣職棒史上最強左打者之一，兩聯盟合計逾 200 支全壘打，是本屆最具攻擊力的外野砲手。",
    ("Wei-Chu Lin",       2006): "2006 WBC 台灣代表隊左外野手，旅日效力阪神虎 10 年，2007 年單季擊出 15 支全壘打，是當時台灣旅日球員中最受期待的攻擊型外野手。",
    ("Chih-Yao Chan",     2006): "2006 WBC 台灣代表隊外野手，中職生涯五度榮獲金手套獎，守備範圍廣闊被譽為「中外野藝術家」，2012 年台灣大賽的飛撲接球成為中職永恆經典畫面。",
    ("Lung-Yi Huang",     2006): "2006 WBC 台灣代表隊外野手，以守備著稱的本土好手，是第一屆世界棒球經典賽台灣外野陣容的重要成員。",
    ("Chien-Ming Chiang", 2006): "2006 WBC 台灣代表隊投手，2006 年以防禦率 1.81 效力讀賣巨人，是當時最受矚目的台灣旅日投手之一。",
    ("Wei-Ming Chu",      2006): "2006 WBC 台灣代表隊投手，新人年繳出防禦率 1.09 的亮眼成績，以牛棚主力身份入選第一屆世界棒球經典賽。",
    ("Wei-Lun Pan",       2006): "2006 WBC 台灣代表隊先發投手，中職生涯累積 149 勝，單季 16 連勝與跨季 21 連勝均為本土紀錄，是中職史上先發出場次數最多的傳奇王牌，2024 年引退。",
    ("En-Yu Lin",         2006): "2006 WBC 台灣代表隊先發投手，首屆 WBC 對韓國首戰投出的開賽第一球被美國棒球名人堂永久典藏，同年更奪下中職勝投、三振、防禦率三冠王。",
    ("Po-Hsuan Keng",     2006): "2006 WBC 台灣代表隊投手，控球精準為其特色，曾旅外效力多倫多藍鳥小聯盟，返台後成為球隊可靠的後援支柱。",
    ("Chu-Chien Hsu",     2006): "2006 WBC 台灣代表隊先發投手，高中時代即以 151 公里強速球揚名，1999 年洲際盃對古巴展現主宰力，是第一屆 WBC 台灣先發陣容的重要成員。",
    ("Wen-Hsiung Hsu",    2006): "2006 WBC 台灣代表隊投手，以獨特的投球姿勢聞名，2006 年亞洲職棒大賽主投六局僅失一分，是本屆陣容中的個性派好手。",
    ("Kevin Huang",       2006): "2006 WBC 台灣代表隊投手，曾以快速球獲波士頓紅襪簽約，為台灣第六位進入美國小聯盟的球員，返台後成為球隊主力後援投手。",
    ("Hung-Chih Kuo",     2006): "2006 WBC 台灣代表隊投手，後成為台灣首位入選 MLB 全明星賽的球員，2010 年以防禦率 1.20 創下道奇隊史單季最低紀錄，大聯盟生涯共登板 218 場。",
    ("Ying-Chieh Lin",    2006): "2006 WBC 台灣代表隊投手，1999 年以 18 歲之齡成為中職史上最年輕先發投手，2004 年奪下三振王與防禦率王，後旅日效力東北樂天金鷲。",
    ("Ying-Feng Tsai",    2006): "2006 WBC 台灣代表隊投手，以快速球為主，是第一屆世界棒球經典賽台灣牛棚的重要成員。",
    ("Sung-Wei Tseng",    2006): "2006 WBC 台灣代表隊唯一業餘球員，出色表現獲克里夫蘭印地安人以 38.5 萬美元簽約，為台灣首位加盟印地安人的選手，後返台兩度克服投球失憶症重返中職。",
    ("Chien-Fu Yang",     2006): "2006 WBC 台灣代表隊投手，以「滑球王子」聞名，2004 年以 15 勝 6 敗、防禦率 1.77 封王並帶領興農牛奪下總冠軍，是中職最具代表性的左投之一。",
    ("Yao-Hsun Yang",     2006): "2006 WBC 台灣代表隊投手，為陽岱鋼之兄，生涯兼任投手與外野手，代表台灣出戰第一屆世界棒球經典賽是其棒球生涯的重要里程碑。",

    # ── 🇹🇼 2013 ──────────────────────────────────────────────────────
    ("Hung-Yu Lin",       2013): "2013 WBC 台灣代表隊主力捕手，2011 年拿下中職年度 MVP，中職生涯累積 216 支全壘打，是桃猿黃金王朝的進攻核心。",
    ("Chih-Kang Kao",     2013): "2013 WBC 台灣代表隊捕手，效力統一獅 14 年，以穩健守備與精準配球著稱，是中職第三位達成千場捕手出賽里程碑的球員。",
    ("Cheng-Min Peng",    2013): "2013 WBC 對澳洲一役以身體擋下界外球再揮出全壘打，「鐵血精神」感動全台；「中職先生」彭政閔生涯 19 年累積 2000+ 安打，持有中職最多保送與得分等多項紀錄。",
    ("Yen-Wen Kuo",       2013): "2013 WBC 台灣代表隊二壘手，曾以自由球員身份簽約辛辛那提紅人並代表台灣出戰多屆 WBC，是早期台灣旅美選手的代表人物。",
    ("Chih-Sheng Lin",    2013): "2013 WBC 台灣代表隊游擊手，雖本職為一壘手仍肩負守備重任，2023 年成為中職首位累計 300 支全壘打的球員。",
    ("Yung-Chi Chen",     2013): "2013 WBC 台灣代表隊三壘手，效力統一獅長達 16 年，曾於 2006 年 WBC 轟出中華隊史上首支滿貫炮，是兩屆經典賽的進攻要角。",
    ("Chiang-Ho Chen",    2013): "2013 WBC 台灣代表隊二壘手，效力中信兄弟逾十年，多次榮獲中職金手套，以精湛的二壘守備著稱於聯盟。",
    ("Han Lin",           2013): "2013 WBC 台灣代表隊內野手，2011 年世界港口賽以打擊率 .538 一舉橫掃最佳打者、最受歡迎球員與大賽 MVP 三獎，是首位獲此殊榮的台灣球員。",
    ("Chien-Ming Chang",  2013): "2013 WBC 台灣代表隊右外野手，以超強臂力聞名的「雷射砲手」，多次代表台灣出征國際大賽，累積中職外野手歷史最多助殺紀錄之一。",
    ("Dai-Kang Yang",     2013): "2013 WBC 奪下 B 組最有價值球員，同年在日職轟出 47 次盜壘稱霸盜壘王並奪得金手套；旅日十餘年累積 105 轟，是台日兩地球迷公認的國民英雄。",
    ("Che-Hsuan Lin",     2013): "2013 WBC 台灣代表隊中外野手，曾以台灣出身球員身份登上波士頓紅襪，為第八位登上大聯盟的台灣球員，2012 年奪下中職台灣大賽 MVP。",
    ("Szu-Chi Chou",      2013): "2013 WBC 台灣代表隊指定打擊，效力中信兄弟 20 年，生涯累積 1804 支安打，是中職史上最具代表性的打擊傳奇，2024 年以四萬人見證的引退典禮完美謝幕。",
    ("Cheng-Wei Chang",   2013): "2013 WBC 台灣代表隊打者，以穩健打擊與替補深度為球隊貢獻板凳火力。",
    ("Yi-Chuan Lin",      2013): "2013 WBC 台灣代表隊打者，2007 年成為中職史上首位同年奪得年度 MVP 與新人王的野手，生涯 2024 年完成中職第 2000 支安打里程碑。",
    ("Chien-Ming Wang",   2013): "2013 WBC 台灣代表隊先發投手，以洋基生涯兩度 19 勝奠定台灣旅美投手最高紀錄；傷後復出以伸卡球帶領台灣首次晉級 WBC 八強，是中華隊精神象徵。",
    ("Wei-Lun Pan",       2013): "2013 WBC 台灣代表隊投手，中職生涯累積 149 勝，單季 16 連勝、跨季 21 連勝均為本土紀錄，是中職史上先發出場次數最多的傳奇王牌。",
    ("Hung-Chih Kuo",     2013): "2013 WBC 台灣代表隊投手，兩度重建韌帶後仍登上大聯盟，2010 年以防禦率 1.20 創道奇隊史最低紀錄，並成為首位入選 MLB 全明星賽的台灣球員。",
    ("Hung-Wen Chen",     2013): "2013 WBC 台灣代表隊終結者，生涯達成中職第五位 100 次救援里程碑，連續出戰 2009、2013、2017 三屆 WBC，是中華隊最忠實的後援靠山。",
    ("Ching-Ming Wang",   2013): "2013 WBC 台灣代表隊投手，曾以日職橫濱球員身份返台效力桃猿，2011 年台灣大賽三連勝奪下 MVP，協助球隊奪冠，是桃猿黃金時期的制勝關鍵。",
    ("Yu-Ching Lin",      2013): "2013 WBC 台灣代表隊投手，WBC 資格賽投出 4 局 5 K 的封關表現助台灣晉級，中職首輪選秀，以三振力聞名聯盟。",
    ("Yi-Hao Lin",        2013): "2013 WBC 台灣代表隊投手，以側投特殊出手角度製造打者不適應，高中時代球速達 149 公里，是中華隊牛棚的個性派成員。",
    ("Jen-Ho Tseng",      2013): "2013 WBC 台灣代表隊最年輕成員之一，同年以非選秀身份加盟芝加哥小熊，2017 年完成大聯盟初登板，是跨越 MLB 與 CPBL 的全方位投手。",
    ("I-Cheng Wang",      2013): "2013 WBC 台灣代表隊投手，旅日橫濱後返台效力桃猿，協助球隊奪下五座中職總冠軍，是桃猿王朝的穩定後援核心。",
    ("Yao-Lin Wang",      2013): "2013 WBC 台灣代表隊投手，為旅美大聯盟球員王維中之兄，代表台灣出戰世界棒球經典賽，是中華隊投手群的組成成員。",

    # ── 🇹🇼 2017 ──────────────────────────────────────────────────────
    ("Kun-Sheng Lin",     2017): "2017 WBC 台灣代表隊主力捕手，效力統一獅 12 年，生涯累積 494 支安打，以穩健配球著稱，曾多次入選中職全明星賽。",
    ("Ta-Hung Cheng",     2017): "2017 WBC 台灣代表隊捕手，是第一屆世界棒球經典賽台灣備援捕手，為球隊提供可靠的後援守備支援。",
    ("Chih-Hsien Chiang", 2017): "2017 WBC 台灣代表隊三壘手，中職以穩健守備與長打能力著稱，是台灣陣容中的主力角落野手。",
    ("Yi-Chuan Lin",      2017): "2017 WBC 台灣代表隊一壘手，連續代表台灣出戰 2013 與 2017 兩屆 WBC，是中職 2000 安成員，生涯橫跨三個十年的中職老將。",
    ("Yung-Chi Chen",     2017): "2017 WBC 台灣代表隊游擊手，統一獅生涯核心球員，在三屆 WBC（2006、2013、2017）均代表台灣出征，是中華隊最資深的野手之一。",
    ("Chi-Hung Hsu",      2017): "2017 WBC 台灣代表隊內野手，以穩定的守備與板凳深度貢獻中華隊，是台灣陣容中不可或缺的工具型球員。",
    ("Chih-Hsiang Lin",   2017): "2017 WBC 台灣代表隊內野手，效力中職以守備扎實著稱，在 2017 年代表台灣出戰世界棒球經典賽。",
    ("Sheng-Wei Wang",    2017): "2017 WBC 台灣代表隊游擊手，以守備穩定著稱的本土球員，為台灣陣容提供游擊守備縱深。",
    ("Chih-Sheng Lin",    2017): "2017 WBC 台灣代表隊球員，生涯以強打見長，中職首位 300 全壘打達成者，年歲漸長仍持續代表國家隊效命。",
    ("Chih-Hao Chang",    2017): "2017 WBC 台灣代表隊左外野手，效力中信兄弟，中職生涯以犀利長打著稱，擁有多項中職三壘安打紀錄，是本屆陣容最具長打威脅的外野砲手。",
    ("Che-Hsuan Lin",     2017): "2017 WBC 台灣代表隊中外野手，連續代表台灣出戰 2013 與 2017 兩屆 WBC，旅日 NPB 期間守備多次獲獎肯定，是中華隊防守最穩固的中堅手。",
    ("Chin-Lung Hu",      2017): "2017 WBC 台灣代表隊指定打擊，曾以台灣首位在 MLB 擊出全壘打的野手身份寫下歷史，返台後創下中職最快達成千安紀錄，是中職守備最全能的游擊傳奇。",
    ("Kuo Hui Lo",        2017): "2017 WBC 台灣代表隊右外野手，以強力長打聞名中職，是台灣陣容中值得對手警戒的重砲外野手。",
    ("Chen-Hua Lin",      2017): "2017 WBC 台灣代表隊投手，以精準控球馳騁中職牛棚，是 2017 年台灣陣容中的重要後援成員。",
    ("Shao-Ching Chiang", 2017): "2017 WBC 台灣代表隊投手，控球精準，旅外期間接受美職系統培訓，返台以後援身份穩定貢獻。",
    ("Hung-Wen Chen",     2017): "2017 WBC 台灣代表隊投手，連續三屆（2009、2013、2017）代表台灣出戰 WBC，是中職第五位達成百次救援的終結者，生涯橫跨三個世代。",
    ("Sheng-Hsiung Huang",2017): "2017 WBC 台灣代表隊投手，2014 年奪下中職三振王（119K），是繼傅力天後本土三振王的繼承者，以高三振率著稱的牛棚要角。",
    ("Fu-Te Ni",          2017): "2017 WBC 台灣代表隊投手，為台灣首位從中職直升大聯盟的本土球員，效力底特律老虎，以豐富的 MLB 經驗為台灣牛棚提供最高等級的國際賽歷練。",
    ("Ming-Chin Tsai",    2017): "2017 WBC 台灣代表隊投手，克服 Tommy John 手術後成功復出，是中職牛棚的穩定後援力量。",
    ("Ching-Ming Wang",   2017): "2017 WBC 台灣代表隊投手，連續代表台灣出戰 2013 與 2017 兩屆 WBC，以 2011 年台灣大賽連拿三勝的表現確立牛棚地位。",
    ("Wei-Lun Pan",       2017): "2017 WBC 台灣代表隊投手，中職史上勝投最多的本土王牌，16 度入選中職全明星賽，在生涯末期仍持續為國家隊效命。",
    ("Chia-Hao Sung",     2017): "2017 WBC 台灣代表隊先發投手，長期旅日效力東北樂天金鷲，是少數在 NPB 一軍穩定出賽的台灣投手，也曾代表台灣出戰 2023 年 WBC。",
    ("Kuo-Hua Lo",        2017): "2017 WBC 台灣代表隊投手，曾旅外接受美職系統培訓，返台後成為中職本土後援選手，是台灣陣容中的板凳後援成員。",

    # ── 🇹🇼 2023 ──────────────────────────────────────────────────────
    ("Kungkuan Giljegiljaw", 2023): "2023 WBC 台灣代表隊主力捕手，以原住民名字為人所知，對義大利一戰敲出關鍵全壘打助台灣大勝，球場上充滿爆發力的攻守表現讓他成為本屆最受矚目的台灣本土捕手。",
    ("Yu-Chieh Kao",      2023): "2023 WBC 台灣代表隊備援捕手，以穩健守備提供球隊後援，是台灣陣容中的板凳戰力。",
    ("Dai-An Lin",        2023): "2023 WBC 台灣代表隊備援捕手，為球隊提供捕手縱深，是台灣陣容中的備援成員。",
    ("Yu Chang",          2023): "2023 WBC Pool A 最有價值球員，全場轟出 3 支全壘打（對義大利 2 發、對荷蘭 1 發），打點高居台灣全隊之冠；MLB 生涯效力克里夫蘭等多隊，以守備精湛的游擊手身份縱橫大聯盟。",
    ("Li Lin",            2023): "2023 WBC 台灣代表隊指定打擊，中職資深打者，以穩定的打擊與豐富的大賽經驗為台灣陣容提供厚度。",
    ("Nien-Ting Wu",      2023): "2023 WBC 台灣代表隊三壘手，效力日職埼玉西武獅長達七年，以守備穩健與接觸打擊著稱，是台灣旅日球員的代表之一。",
    ("Tsung-Che Cheng",   2023): "2023 WBC 台灣代表隊二壘手，以速度與精準打擊聞名中職，是台灣陣容中攻守均衡的核心中線球員。",
    ("Kun-Yu Chiang",     2023): "2023 WBC 台灣代表隊游擊手，中職生涯 6 連座黃金手套，成為中職首位達成此成就的內野手，以生涯最高合約確立最佳游擊手地位。",
    ("Tzu-Wei Lin",       2023): "2023 WBC 台灣代表隊外野手，曾效力波士頓紅襪，是台灣少數在 MLB 以工具型球員身份立足的選手，守備可覆蓋多個位置。",
    ("Kuo-Chen Fan",      2023): "2023 WBC 台灣代表隊一壘手，以本土好手身份為台灣陣容提供一壘守備深度。",
    ("Wei-Chen Wang",     2023): "2023 WBC 台灣代表隊三壘手，中職以穩健守備見長，為台灣角落野手陣容貢獻守備實力。",
    ("Chieh-Hsien Chen",  2023): "2023 WBC 台灣代表隊中外野手，4 連座 CPBL 金手套外野，2019 年 Premier12 大賽 MVP，是當代台灣最全面的中外野手，攻守均是球隊主軸。",
    ("Po-Jung Wang",      2023): "2023 WBC 台灣代表隊外野手，中職史上打擊成就最高的球員之一，2016 年以 .414 打擊率寫下聯盟歷史，2017 年奪下四冠王，是中職黃金世代的代表人物。",
    ("Chin Cheng",        2023): "2023 WBC 台灣代表隊外野手，以穩定的中職表現入選，為台灣陣容提供外野板凳深度。",
    ("Tien-Hsin Kuo",     2023): "2023 WBC 台灣代表隊外野手，中職本土球員，以守備與接觸打擊貢獻台灣陣容。",
    ("Chia-Hao Sung",     2023): "2023 WBC 台灣代表隊終結者，長期效力日職東北樂天金鷲，是少數在 NPB 一軍穩定出賽的台灣投手，連續代表台灣出戰 2017 與 2023 兩屆 WBC。",
    ("Che-Yuan Wu",       2023): "2023 WBC 台灣代表隊後援投手，中職主力後援，以控球精準為牛棚貢獻穩定的出賽品質。",
    ("Yen-Ching Lu",      2023): "2023 WBC 台灣代表隊後援投手，以犀利的球路組合馳騁中職牛棚，是台灣陣容中的重要後援成員。",
    ("Wei-Chung Wang",    2023): "2023 WBC 台灣代表隊後援投手，以左投身份對抗左打者見長，是台灣牛棚中的左投剋星利器。",
    ("Yu-Hsun Chen",      2023): "2023 WBC 台灣代表隊後援投手，中職以穩定後援表現著稱，是台灣陣容牛棚的可靠支援。",
    ("Tzu-Peng Huang",    2023): "2023 WBC 台灣代表隊先發投手，以中職先發身份入選，為球隊提供先發陣容縱深。",
    ("Kuan-Yu Chen",      2023): "2023 WBC 台灣代表隊投手，效力日職橫濱 DeNA 海灣之星與千葉羅德海洋馬林魚合計逾十年，是台灣旅日投手中在 NPB 一軍資歷最深的球員之一。",
    ("Shih-Peng Chen",    2023): "2023 WBC 台灣代表隊先發投手，中職本土好手，以先發穩定表現入選中華隊，為台灣投手陣容盡力。",
    ("C.C. Lee",          2023): "2023 WBC 台灣代表隊後援投手，以中職主力後援身份入選，為台灣牛棚貢獻出賽實力。",
    ("Kai-Wei Teng",      2023): "2023 WBC 台灣代表隊後援投手，中職以速球為主的右投後援，是台灣陣容牛棚的重要成員。",
    ("Jyun-Yue Tseng",    2023): "2023 WBC 台灣代表隊後援投手，中職以犀利球路見長，本屆擔任牛棚後援角色，後成長為中華隊主力終結者。",
    ("Kuan-Wei Chen",     2023): "2023 WBC 台灣代表隊後援投手，中職本土後援球員，為台灣陣容提供牛棚縱深。",
    ("Chih-Wei Hu",       2023): "2023 WBC 台灣代表隊先發投手，中職以穩健先發身份入選，為球隊貢獻先發出賽能量。",

    # ── 🇹🇼 2026 ──────────────────────────────────────────────────────
    ("Lyle Lin",          2026): "2026 WBC 台灣代表隊捕手，以穩健守備與精準配球著稱，是新生代台灣本土捕手的代表人物。",
    ("Shao-Hung Chiang",  2026): "2026 WBC 台灣代表隊備援捕手，為球隊提供捕手守備縱深。",
    ("Kungkuan Giljegiljaw", 2026): "2026 WBC 台灣代表隊，曾在 2023 年 WBC 以全壘打震撼國際舞台，本屆轉任指定打擊繼續為台灣效命。",
    ("Tsung-Che Cheng",   2026): "2026 WBC 台灣代表隊核心二壘手，中職以速度與精準打擊聞名，已成為台灣陣容中最不可或缺的中線攻擊核心。",
    ("Yu Chang",          2026): "2026 WBC 台灣代表隊三壘手，2023 年 WBC Pool A MVP 得主，連續兩屆代表台灣出征，以 MLB 實戰經驗帶領中華隊面對世界最強對手。",
    ("Cheng-Yu Chang",    2026): "2026 WBC 台灣代表隊三壘手，以本土好手身份入選，為台灣角落野手陣容提供守備深度。",
    ("Kun-Yu Chiang",     2026): "2026 WBC 台灣代表隊游擊手，中職六連座黃金手套的守備傳奇，連續代表台灣出戰 2023 與 2026 兩屆 WBC，是中華隊游擊守備無可撼動的第一人。",
    ("Tzu-Wei Lin",       2026): "2026 WBC 台灣代表隊球員，工具型好手，可覆蓋內外野多個位置，豐富的旅外經驗使其在高強度國際賽中依然保持穩定。",
    ("Nien-Ting Wu",      2026): "2026 WBC 台灣代表隊一壘手，旅日七年效力埼玉西武獅後返台，2024 年以中職首輪選秀加盟 TSG 鷹，打擊率 .304 再現往日水準。",
    ("Stuart Fairchild",  2026): "2026 WBC 以母親的台灣血統代表中華隊出賽，對捷克一戰轟出滿貫全壘打，全場打擊率 .250 並擊出 2 支全壘打，速度與守備出眾的中外野手令台灣球迷眼睛一亮。",
    ("Chen-Wei Chen",     2026): "2026 WBC 台灣代表隊左外野手，中職以穩健打擊著稱，是台灣外野陣容的組成成員。",
    ("Chieh-Hsien Chen",  2026): "2026 WBC 台灣代表隊外野手，4 連座 CPBL 金手套外野、2019 Premier12 MVP，連續代表台灣出戰 2023 與 2026 兩屆 WBC，是中華隊外野守備最堅實的保障。",
    ("An-Ko Lin",         2026): "2026 WBC 台灣代表隊右外野手，以速度與守備著稱的本土好手，為台灣外野陣容提供機動深度。",
    ("Cheng-Hui Sung",    2026): "2026 WBC 台灣代表隊右外野手，中職以接觸打擊與速度見長，是台灣陣容的機動外野選項。",
    ("Jyun-Yue Tseng",    2026): "2026 WBC 台灣代表隊終結者，從 2023 年的後援成員成長為本屆主力封關投手，以犀利球路在中職確立終結者地位。",
    ("Yu-Min Lin",        2026): "2026 WBC 台灣代表隊後援投手，以速球與變化球組合馳騁中職牛棚，是台灣陣容重要的後援核心。",
    ("Chen Zhong-Ao Zhuang", 2026): "2026 WBC 台灣代表隊先發投手，以中職先發穩定表現入選，為台灣投手陣容提供先發輪值深度。",
    ("Jo-Hsi Hsu",        2026): "2026 WBC 台灣代表隊先發投手，以精準控球與多樣球路在中職闖出名號，是台灣新生代先發王牌的代表人物。",
    ("Ruei-Yang Gu Lin",  2026): "2026 WBC 台灣代表隊先發投手，2024 年以 ERA 1.66、10 勝奪下 CPBL 年度 MVP，成為中職 18 年來首位本土 MVP 先發投手，同年轉赴日職北海道日本火腿鬥士，在 NPB 投出完封勝（Maddux），是台灣目前最頂尖的旅外投手。",
    ("Wei-En Lin",        2026): "2026 WBC 台灣代表隊後援投手，以中職主力後援身份入選，為台灣牛棚提供穩定的後援戰力。",
    ("Jun-Wei Zhang",     2026): "2026 WBC 台灣代表隊後援投手，以速球為主的中職右投後援，是台灣陣容牛棚組成成員。",
    ("Yi-Lei Sun",        2026): "2026 WBC 台灣代表隊後援投手，中職以穩定後援表現獲選，為台灣陣容牛棚盡力。",
    ("Yi Chang",          2026): "2026 WBC 台灣代表隊後援投手，以犀利球路在中職牛棚站穩腳跟，是台灣新生代後援的代表人物。",
    ("Kuan-Yu Chen",      2026): "2026 WBC 台灣代表隊投手，旅日逾十年，以豐富的 NPB 經驗繼續為中華隊投手群貢獻國際賽歷練。",
    ("Po-Yu Chen",        2026): "2026 WBC 台灣代表隊後援投手，以中職本土身份入選，為台灣陣容牛棚提供縱深。",
    ("Hao-Chun Cheng",    2026): "2026 WBC 台灣代表隊先發投手，中職以先發穩定性著稱，是台灣投手輪值的組成成員。",
    ("Chih-Wei Hu",       2026): "2026 WBC 台灣代表隊投手，連續代表台灣出戰 2023 與 2026 兩屆 WBC，以中職穩定表現維繫在國家隊的地位。",
    ("Kai-Wei Lin",       2026): "2026 WBC 台灣代表隊後援投手，以速球主攻的中職本土後援，是台灣牛棚的成員之一。",
    ("Shih-Hsiang Lin",   2026): "2026 WBC 台灣代表隊後援投手，以中職穩定後援身份入選，為台灣陣容貢獻牛棚深度。",
    ("Tzu-Chen Sha",      2026): "2026 WBC 台灣代表隊後援投手，以犀利球路在中職牛棚累積出賽資歷，是台灣陣容的後援成員。",
}

# ── story ─────────────────────────────────────────────────────────────────────

def make_story(d: dict, is_pitcher: bool) -> str:
    sea = d["season"]
    nat = d["nat"]
    ovr = safe_int(d["RTG_OVR"])
    pos = d["pos"] or ""
    pz  = POS_ZH.get(pos, pos)
    en  = d.get("name_en") or ""

    # Year-specific override first, then name-only fallback
    story_key = (en, int(sea)) if (en, int(sea)) in STORY_OVERRIDES else en
    if story_key in STORY_OVERRIDES:
        return STORY_OVERRIDES[story_key]

    if is_pitcher:
        ctrl = safe_int(d["RTG_CTRL"])
        k    = safe_int(d["RTG_K"])
        if ovr >= 88:
            feat = "強力三振與精準控球" if k > ctrl else "壓倒性球威與穩定表現"
            return f"{sea} WBC {nat}代表隊頂尖{pz}，以{feat}主宰國際賽場，是當屆最強投手之一。"
        elif ovr >= 78:
            return f"{sea} WBC {nat}代表隊主力{pz}，穩定的投球表現是球隊投手群的重要支柱。"
        else:
            return f"{sea} WBC {nat}代表隊{pz}，在世界棒球經典賽中貢獻出賽實力。"
    else:
        pwr = safe_int(d["RTG_POWER"])
        spd = safe_int(d["RTG_SPEED"])
        if ovr >= 90:
            if pwr >= 85:
                feat = "驚人的長打威脅力"
            elif spd >= 85:
                feat = "速度與打擊的完美結合"
            else:
                feat = "全方位頂尛打擊實力"
            return f"{sea} WBC {nat}代表隊核心{pz}，以{feat}震撼國際舞台，是當屆最耀眼的球員之一。"
        elif ovr >= 80:
            return f"{sea} WBC {nat}代表隊主力{pz}，打擊表現出色，是球隊進攻火力的重要支柱。"
        else:
            return f"{sea} WBC {nat}代表隊{pz}，在世界棒球經典賽中展現了實力。"


# ── entry ─────────────────────────────────────────────────────────────────────

def make_entry(d: dict) -> str:
    pos = (d["pos"] or "OF").strip()
    is_p = pos in PITCHER_POS

    name_zh = (d["name_zh"] or "").strip()
    name_en = (d["name_en"] or "").strip()
    display  = name_zh if name_zh else name_en

    nat   = d["nat"]
    flag  = NAT_TO_FLAG.get(nat, "🌐")
    ovr   = safe_int(d["RTG_OVR"])
    sea   = d["season"]

    # Apply OVR floor for world-class players with small/unlucky WBC samples
    floor = OVR_FLOOR.get(name_en, 0)
    if ovr < floor:
        ovr = floor
        d["RTG_OVR"] = floor   # propagate to bat_stats / pit_stats / make_story

    # Apply OVR ceiling for players whose WBC small sample inflated their rating
    ceiling = OVR_CEILING.get(name_en, 99)
    if ovr > ceiling:
        ovr = ceiling
        d["RTG_OVR"] = ceiling

    # Apply global shift (compensates for WBC short-sample bias); cap at 99
    ovr = min(99, ovr + OVR_SHIFT)
    d["RTG_OVR"] = ovr

    stats_val  = pit_stats(d) if is_p else bat_stats(d)
    power_val  = None if is_p else bat_power(d)
    excel_val  = excel_stats(d, is_p)
    skills_val = make_skills(d, is_p)
    story_val  = make_story(d, is_p)

    def sk(s):
        return "{" + f"i:{js_str(s['i'])},n:{js_str(s['n'])},d:{js_str(s['d'])}" + "}"

    # pos: single string or array for compound
    if "/" in pos:
        parts_pos = pos.split("/")
        pos_js = "[" + ",".join(js_str(p) for p in parts_pos) + "]"
    else:
        pos_js = js_str(pos)

    stats_js  = "[" + ",".join(str(x) for x in stats_val) + "]"
    skills_js = "[" + ",".join(sk(s) for s in skills_val) + "]"

    fields = [
        f"name:{js_str(display)}",
        f"year:{sea}",
        f"en:{js_str(name_en)}",
        f"nat:{js_str(flag)}",
        f"pos:{pos_js}",
        f"av:{js_str(POS_AV.get(pos, '⚾'))}",
        f"rar:{js_str(ovr_to_rar(ovr))}",
        f"ovr:{ovr}",
        f"pit:{'true' if is_p else 'false'}",
        f"era:[{sea}]",
        f"stats:{stats_js}",
        *( [f"power:{power_val}"] if power_val is not None else [] ),
        f"excel:{js_obj(excel_val)}",
        f"story:{js_str(story_val)}",
        f"skills:{skills_js}",
    ]
    return "{" + ",".join(fields) + "}"


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = load_workbook(XLSX)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    RTG = ["RTG_CONTACT","RTG_POWER","RTG_EYE","RTG_SPEED","RTG_BAT",
           "RTG_CTRL","RTG_K","RTG_ERA_Q","RTG_PIT","RTG_OVR"]
    RAW_STATS = [src for src, _ in BAT_EXCEL_FIELDS + PIT_EXCEL_FIELDS]

    def g(r, c):
        return ws.cell(r, col[c]).value if c in col else None

    # Read all rows
    by_nat: dict[str, list[dict]] = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        nat = g(r, "nationality")
        if nat not in NAT_TO_FILE:
            continue
        d = {
            "nat":     nat,
            "season":  g(r, "season"),
            "section": g(r, "section") or "Pitchers",
            "name_en": g(r, "player_name"),
            "name_zh": g(r, "player_name_zh"),
            "pos":     g(r, "POS") or "OF",
            **{rc: g(r, rc) for rc in RTG},
            **{rc: g(r, rc) for rc in RAW_STATS},
        }
        by_nat[nat].append(d)

    lines: list[str] = []
    lines.append("/* ============================================================")
    lines.append("   Diamond Nations — players-all.js")
    lines.append(f"   球員資料庫：全 {sum(len(v) for v in by_nat.values()) + len(REGULAR_EXTRAS) + len(LEGENDS)} 筆（由 wbc_roster.xlsx 生成）")
    lines.append("   ============================================================ */")
    lines.append("")
    lines.append("window.RAW_ALL_PLAYERS=[")

    for nat in NAT_TO_FILE:
        rows = by_nat.get(nat, [])
        if not rows:
            continue
        disp = NAT_DISPLAY.get(nat, nat)

        # Group by section
        by_sec: dict[str, list[dict]] = defaultdict(list)
        for d in rows:
            sec = d["section"]
            if sec not in SECTION_ZH:
                pos = d["pos"] or ""
                if pos in PITCHER_POS:       sec = "Pitchers"
                elif pos == "C":             sec = "Catchers"
                elif pos == "DH":            sec = "Designated Hitter"
                elif pos in ("LF","CF","RF","OF"): sec = "Outfielders"
                else:                        sec = "Infielders"
            by_sec[sec].append(d)

        for sec in by_sec:
            by_sec[sec].sort(key=lambda x: (x["season"], -safe_int(x["RTG_OVR"])))

        lines.append(f"")
        lines.append(f"/* {'═'*36}")
        lines.append(f"   {disp}  ({len(rows)} 人)")
        lines.append(f"{'═'*36} */")

        for sec in SECTION_ORDER:
            if sec not in by_sec:
                continue
            lines.append(f"/* ── {SECTION_ZH.get(sec, sec)} ── */")
            for d in by_sec[sec]:
                lines.append(make_entry(d) + ",")

    # ── 傳奇球員一般版（補上無 WBC 記錄的傳奇球員一般版）────────────────────────
    lines.append("")
    lines.append(f"/* {'═'*36}")
    lines.append("   🃏 傳奇球員一般版（補充卡）")
    lines.append(f"{'═'*36} */")
    for entry in REGULAR_EXTRAS:
        lines.append(entry + ",")

    # ── 傳奇球員（手動，rar:'x'，ovr:99）────────────────────────────────────────
    lines.append("")
    lines.append(f"/* {'═'*36}")
    lines.append("   📸 復古傳奇特別款（全 8 國）")
    lines.append(f"{'═'*36} */")
    for entry in LEGENDS:
        lines.append(entry + ",")

    lines.append("")
    lines.append("];")
    lines.append("")
    lines.append("/* 由 players-data.js 接手正規化與驗證 */")
    lines.append("")

    OUTPUT.write_text("\n".join(lines), encoding="utf-8")
    total = sum(len(v) for v in by_nat.values()) + len(REGULAR_EXTRAS) + len(LEGENDS)
    size  = OUTPUT.stat().st_size / 1024
    print(f"✅ {OUTPUT}  ({total} players, {size:.0f} KB)")
    print("Done.")


if __name__ == "__main__":
    main()
