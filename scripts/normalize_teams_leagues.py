"""
Normalize team (→ Chinese) and league (→ short code) columns in Sheet1.
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

OUTPUT = Path(__file__).parent.parent / "wbc_roster.xlsx"

# ── League normalization ────────────────────────────────────────────────────
LEAGUE_MAP: dict[str, str] = {
    # MLB / MiLB
    "Major League Baseball":            "MLB",
    "MLB":                              "MLB",
    "Minor League Baseball":            "MiLB",
    "MiLB":                             "MiLB",
    "Independent":                      "IND",
    # CPBL
    "CPBL":                             "CPBL",
    "Chinese Professional Baseball":    "CPBL",
    # NPB
    "Nippon Professional Baseball":     "NPB",
    "NPB":                              "NPB",
    "NPB Farm":                         "NPB",
    # KBO
    "KBO League":                       "KBO",
    "KBO":                              "KBO",
    # Cuba
    "Cuban National Series":            "CNS",
    "Cuban National Series (Serie Nacional)": "CNS",
    # Venezuela
    "Liga Venezolana de Béisbol Profesional": "LVBP",
    "LVBP":                             "LVBP",
    # Mexico
    "Mexican League":                   "LMB",
    # Puerto Rico
    "Liga Roberto Clemente":            "LRC",
    # Misc
    "Amateur":                          "業餘",
    "無":                               "",
    # Cuban city/team values that ended up in league column
    "Villa Clara":                      "CNS",
    "Antilla":                          "CNS",
    "Guayama":                          "LRC",
    "Mayagüez":                         "LRC",
    "Naguabo":                          "LRC",
    "Cayey":                            "LRC",
}

# ── Team name → Chinese ────────────────────────────────────────────────────
TEAM_MAP: dict[str, str] = {
    # ── MLB (30 teams) ──
    "Arizona Diamondbacks":     "亞利桑那響尾蛇",
    "Atlanta Braves":           "亞特蘭大勇士",
    "Baltimore Orioles":        "巴爾的摩金鶯",
    "Boston Red Sox":           "波士頓紅襪",
    "Chicago Cubs":             "芝加哥小熊",
    "Chicago White Sox":        "芝加哥白襪",
    "Cincinnati Reds":          "辛辛那提紅人",
    "Cleveland Indians":        "克里夫蘭印地安人",
    "Cleveland Guardians":      "克里夫蘭守護者",
    "Colorado Rockies":         "科羅拉多洛磯",
    "Detroit Tigers":           "底特律老虎",
    "Houston Astros":           "休士頓太空人",
    "Kansas City Royals":       "堪薩斯城皇家",
    "Los Angeles Angels":       "洛杉磯天使",
    "Los Angeles Dodgers":      "洛杉磯道奇",
    "Miami Marlins":            "邁阿密馬林魚",
    "Milwaukee Brewers":        "密爾瓦基釀酒人",
    "Minnesota Twins":          "明尼蘇達雙城",
    "New York Mets":            "紐約大都會",
    "New York Yankees":         "紐約洋基",
    "Oakland Athletics":        "奧克蘭運動家",
    "Athletics":                "奧克蘭運動家",
    "Philadelphia Phillies":    "費城費城人",
    "Pittsburgh Pirates":       "匹茲堡海盜",
    "San Diego Padres":         "聖地牙哥教士",
    "San Francisco Giants":     "舊金山巨人",
    "Seattle Mariners":         "西雅圖水手",
    "St. Louis Cardinals":      "聖路易紅雀",
    "Tampa Bay Rays":           "坦帕灣光芒",
    "Texas Rangers":            "德克薩斯遊騎兵",
    "Toronto Blue Jays":        "多倫多藍鳥",
    "Washington Nationals":     "華盛頓國民",
    # ── NPB ──
    "Fukuoka SoftBank Hawks":           "福岡軟銀鷹",
    "Yomiuri Giants":                   "讀賣巨人",
    "Chunichi Dragons":                 "中日龍",
    "Chiba Lotte Marines":              "千葉羅德海洋",
    "Saitama Seibu Lions":              "埼玉西武獅",
    "Orix Buffaloes":                   "歐力士野牛",
    "Tokyo Yakult Swallows":            "東京養樂多燕子",
    "Hokkaido Nippon-Ham Fighters":     "北海道日本火腿鬥士",
    "Tohoku Rakuten Golden Eagles":     "東北樂天金鷹",
    "Hanshin Tigers":                   "阪神虎",
    "Hiroshima Toyo Carp":              "廣島東洋鯉魚",
    "DeNA BayStars":                    "橫濱DeNA海灣之星",
    "Yokohama BayStars":                "橫濱灣之星",
    "Hiroshima Carp":                   "廣島東洋鯉魚",
    "Rakuten Eagles":                   "東北樂天金鷹",
    "Seibu Lions":                      "埼玉西武獅",
    "Nippon-Ham Fighters":              "北海道日本火腿鬥士",
    # ── KBO ──
    "Doosan Bears":             "斗山熊",
    "LG Twins":                 "LG雙子",
    "Samsung Lions":            "三星獅",
    "Lotte Giants":             "樂天巨人",
    "KIA Tigers":               "KIA虎",
    "SK Wyverns":               "SK飛龍",
    "SSG Landers":              "SSG登陸者",
    "NC Dinos":                 "NC恐龍",
    "Kiwoom Heroes":            "奇偉英雄",
    "Nexen Heroes":             "耐克森英雄",
    "Hanwha Eagles":            "漢化鷹",
    "KT Wiz":                   "KT巫師",
    "Kt Wiz":                   "KT巫師",
    "Lotte Giants KBO":         "樂天巨人",
    # ── CPBL (already Chinese, keep as-is — listed for completeness) ──
    "富邦悍將":     "富邦悍將",
    "統一獅":       "統一獅",
    "中信兄弟":     "中信兄弟",
    "樂天桃猿":     "樂天桃猿",
    "味全龍":       "味全龍",
    "三商虎":       "三商虎",
    "兄弟象":       "兄弟象",
    "La New熊":     "La New熊",
    # ── Cuban teams ──
    "Villa Clara":              "比亞克拉拉",
    "Granma":                   "格拉瑪",
    "Pinar del Río":            "比那爾德里奧",
    "Industriales":             "工業家",
    "Holguín":                  "奧爾金",
    "Santiago de Cuba":         "古巴聖地亞哥",
    "Matanzas":                 "馬坦薩斯",
    "Las Tunas":                "拉斯圖納斯",
    "Camagüey":                 "卡馬圭",
    "Ciego de Ávila":           "謝戈德阿維拉",
    "Cienfuegos":               "西恩富戈斯",
    "Sancti Spíritus":          "桑克蒂斯皮里圖斯",
    "Artemisa":                 "阿爾特米薩",
    "Mayabeque":                "馬亞貝克",
    "Isla de la Juventud":      "青年島",
    "Guantánamo":               "關塔那摩",
    "Antilla":                  "安提亞隊",
    # ── Country-code placeholders ──
    "KOR":  "韓國（未記錄）",
    "VEN":  "委內瑞拉（未記錄）",
    "JPN":  "日本（未記錄）",
    "CUB":  "古巴（未記錄）",
    "PUR":  "波多黎各（未記錄）",
    "DOM":  "多明尼加（未記錄）",
    "USA":  "美國（未記錄）",
    # ── Free agent / misc ──
    "Free agent":               "自由球員",
    "Retired":                  "退役",
    # ── CPBL historical names ──
    "Rakuten Monkeys":          "樂天桃猿",
    "Lamigo桃猿":               "樂天桃猿",
    "CTBC Brothers":            "中信兄弟",
    "Fubon Guardians":          "富邦悍將",
    "Uni-President 7-Eleven Lions": "統一獅",
    "Uni-President Lions":      "統一獅",
    "Sinon Bulls":              "興農牛",
    "Wei Chuan Dragons":        "味全龍",
    "Wei Chuan":                "味全龍",
    "La New Bears":             "La New熊",
    "La New熊":                 "La New熊",
    "Macoto Cobras":            "誠泰眼鏡蛇",
    "Brother Elephants":        "兄弟象",
    "Mega Electronics Elephants": "大象隊",
    # ── MLB alternate/old names ──
    "Los Angeles Angels of Anaheim":    "洛杉磯天使",
    "Anaheim Angels":                   "洛杉磯天使",
    "Florida Marlins":                  "佛羅里達馬林魚",
    "Montreal Expos":                   "蒙特婁博覽會",
    "Tampa Bay Devil Rays":             "坦帕灣光芒",
    # ── MLB minor league (keep MiLB tag) ──
    "Cincinnati Reds (minors)":         "辛辛那提紅人（小聯盟）",
    "Minnesota Twins (minors)":         "明尼蘇達雙城（小聯盟）",
    "Miami Marlins (minors)":           "邁阿密馬林魚（小聯盟）",
    "Detroit Tigers (minors)":          "底特律老虎（小聯盟）",
    "Chicago Cubs (minors)":            "芝加哥小熊（小聯盟）",
    "Pittsburgh Pirates (minors)":      "匹茲堡海盜（小聯盟）",
    "Seattle Mariners (minors)":        "西雅圖水手（小聯盟）",
    "Binghamton Mets":                  "賓漢頓大都會（小聯盟）",
    "College of Physical Education":    "體育學院",
    # ── NPB alternate ──
    "Hokkaido Nippon Ham Fighters":     "北海道日本火腿鬥士",
    "Yokohama DeNA BayStars":           "橫濱DeNA海灣之星",
    # ── KBO alternate ──
    "Kia Tigers":               "KIA虎",
    "Seoul Heroes":             "首爾英雄",
    # ── Cuban teams ──
    "Habana":                   "哈瓦那",
    "Cocodrilos de Matanzas":   "馬坦薩斯鱷魚",
    "Vegueros de Pinar del Río":"比那爾德里奧農民",
    "Leñadores de Las Tunas":   "拉斯圖納斯伐木工",
    "Toros de Camagüey":        "卡馬圭公牛",
    # ── Mexican League ──
    "Leones de Yucatán":                "猶加敦雄獅",
    "Algodoneros de Unión Laguna":      "聯合拉古納棉農",
    "Diablos Rojos del México":         "墨西哥紅魔",
    "Sultanes de Monterrey":            "蒙特雷蘇丹",
    "Tigres de Quintana Roo":           "金塔納羅奧虎",
    "Broncos de Reynosa":               "雷諾薩野馬",
    "Pericos de Puebla":                "普埃布拉長尾鸚鵡",
    "Guerreros de Oaxaca":              "瓦哈卡戰士",
    "Olmecas de Tabasco":               "塔巴斯科奧爾梅克",
    "Rieleros de Aguascalientes":       "阿瓜斯卡連特斯鐵路工",
    "Acereros de Monclova":             "蒙克洛瓦鋼鐵工",
    "Saraperos de Saltillo":            "薩爾提約農工",
    "Piratas de Campeche":              "坎佩切海盜",
    "Toros de Tijuana":                 "蒂華納公牛",
    # ── Venezuelan/Puerto Rico teams ──
    "Tiburones de La Guaira":           "拉瓜伊拉鯊魚",
    "Leones del Caracas":               "卡拉卡斯雄獅",
    "Navegantes del Magallanes":        "麥哲倫航海者",
    "Cardenales de Lara":               "拉臘紅雀",
    "Aguilas del Zulia":                "蘇利亞雄鷹",
    "Tigres de Aragua":                 "阿拉瓜虎",
    "Bravos de Margarita":              "馬格麗塔勇士",
    "Caribes de Anzoátegui":            "安索阿特吉加勒比",
    "Gigantes de Caguas":               "卡瓜斯巨人",
    "Criollos de Caguas":               "卡瓜斯克里奧爾",
    "Indios de Mayagüez":               "馬亞圭斯印地安人",
    "Leones de Ponce":                  "彭斯雄獅",
    "Cangrejeros de Santurce":          "聖圖爾塞螃蟹",
    "Senadores de San Juan":            "聖胡安參議員",
    # ── Misc / historical ──
    "Québec Capitales":                 "魁北克首都",
    "La Habana":                        "哈瓦那",
    "Alazanes de Granma":               "格拉瑪飛馬",
    "Cazadores de Artemisa":            "阿爾特米薩獵人",
    "Tigres del Licey":                 "利賽虎",
    "Elefantes de Cienfuegos":          "西恩富戈斯象",
    "Chinatrust Whales":                "中信鯨",
    "Hyundai Unicorns":                 "現代獨角獸",
    "Ducks":                            "鴨隊",
    "Dunedin Blue Jays":                "達尼丁藍鳥（小聯盟）",
    "Inland Empire 66ers":              "內陸帝國66人（小聯盟）",
    "Jacksonville Suns":                "傑克遜維爾太陽（小聯盟）",
    "Free Agent":                       "自由球員",
    "Hokkaido Nippon-Ham Fighters (farm)": "北海道日本火腿鬥士（二軍）",
    "Lamigo Monkeys":                   "樂天桃猿",
    "Aguilas Cibaeñas":                 "錫拜奧雄鷹",
    "Ganaderos de Camagüey":            "卡馬圭牧牛人",
    "Indios de Guantánamo":             "關塔那摩印地安人",
    "Caliente de Durango":              "杜蘭戈火熱",
    "Gigantes de Carolina":             "卡羅萊納巨人",
    "Texas Rangers/Washington Nationals": "德克薩斯遊騎兵/華盛頓國民",
    "Altoona Curve":                    "阿爾圖納曲球（小聯盟）",
    "West Tenn Diamond Jaxx":           "西田納西鑽石賈克斯（小聯盟）",
    "Newark Bears":                     "紐瓦克熊（獨立聯盟）",
}


def normalize_league(val: str) -> str:
    if not val:
        return val
    return LEAGUE_MAP.get(val.strip(), val.strip())


def translate_team(val: str) -> str:
    if not val:
        return val
    v = val.strip()
    if v in TEAM_MAP:
        return TEAM_MAP[v]
    # Handle "TeamName (minors)" / "(farm)" / "(minor league)" variants
    import re
    m = re.match(r"^(.+?)\s*\((minors?|farm|minor league)\)$", v, re.IGNORECASE)
    if m:
        base = translate_team(m.group(1).strip())
        return f"{base}（小聯盟）"
    return v


def main() -> None:
    wb = load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    league_changed = team_changed = 0

    for r in range(2, ws.max_row + 1):
        # League
        cell_l = ws.cell(r, col["league"])
        new_l = normalize_league(cell_l.value or "")
        if new_l != (cell_l.value or ""):
            cell_l.value = new_l
            league_changed += 1

        # Team
        cell_t = ws.cell(r, col["team"])
        new_t = translate_team(cell_t.value or "")
        if new_t != (cell_t.value or ""):
            cell_t.value = new_t
            team_changed += 1

    wb.save(OUTPUT)
    print(f"League normalized: {league_changed}  Team translated: {team_changed}")

    # Report anything still untranslated
    from collections import Counter
    remaining_leagues: Counter = Counter()
    remaining_teams: Counter = Counter()
    for r in range(2, ws.max_row + 1):
        l = ws.cell(r, col["league"]).value
        t = ws.cell(r, col["team"]).value
        if l and not any(ord(c) > 127 or c.isupper() and len(l) <= 4 for c in l):
            # crude check: long English string still there
            if any(c.islower() for c in str(l)) and len(str(l)) > 4:
                remaining_leagues[l] += 1
        if t and any(c.islower() for c in str(t)) and len(str(t)) > 4:
            remaining_teams[t] += 1

    if remaining_leagues:
        print("\nLeagues still in English:")
        for k, v in remaining_leagues.most_common(20):
            print(f"  {v:4d}  {k!r}")
    if remaining_teams:
        print("\nTeams still in English (top 30):")
        for k, v in remaining_teams.most_common(30):
            print(f"  {v:4d}  {k!r}")


if __name__ == "__main__":
    main()
