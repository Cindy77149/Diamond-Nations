"""
Scrape WBC roster tables from Wikipedia and fill team/league columns in Sheet1.

Wikipedia pages used:
  https://en.wikipedia.org/wiki/20XX_World_Baseball_Classic_rosters
"""
from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, Page

OUTPUT = Path(__file__).parent.parent / "wbc_roster.xlsx"

WIKI_URLS = {
    2023: "https://en.wikipedia.org/wiki/2023_World_Baseball_Classic_rosters",
    2017: "https://en.wikipedia.org/wiki/2017_World_Baseball_Classic_rosters",
    2013: "https://en.wikipedia.org/wiki/2013_World_Baseball_Classic_rosters",
    2009: "https://en.wikipedia.org/wiki/2009_World_Baseball_Classic_rosters",
    2006: "https://en.wikipedia.org/wiki/2006_World_Baseball_Classic_rosters",
}

# Wikipedia country name → our nationality key
COUNTRY_MAP = {
    "Taiwan":              "台灣",
    "Chinese Taipei":      "台灣",
    "Japan":               "日本",
    "South Korea":         "韓國",
    "Korea":               "韓國",
    "United States":       "美國",
    "Dominican Republic":  "多明尼加",
    "Venezuela":           "委內瑞拉",
    "Cuba":                "古巴",
    "Puerto Rico":         "波多黎各",
}

EXTRACT_JS = """
() => {
    const clean = s => (s || '').replace(/\\[.*?\\]/g, '').replace(/\\s+/g, ' ').trim();

    const extractName = (cell) => {
        // Try .fn span first (vcard format)
        const fn = cell.querySelector('.fn');
        if (fn) return clean(fn.innerText);
        // Try first <a> link
        const a = cell.querySelector('a');
        if (a) return clean(a.innerText);
        return clean(cell.innerText);
    };

    const items = [];
    const walker = document.createTreeWalker(
        document.body, NodeFilter.SHOW_ELEMENT,
        { acceptNode: n => {
            if (/^H[2-4]$/.test(n.tagName)) return NodeFilter.FILTER_ACCEPT;
            if (n.tagName === 'TABLE' && n.className.includes('wikitable')) return NodeFilter.FILTER_ACCEPT;
            return NodeFilter.FILTER_SKIP;
        }}
    );
    let node;
    while ((node = walker.nextNode())) {
        if (/^H[2-4]$/.test(node.tagName)) {
            items.push({ type: 'heading', text: clean(node.innerText) });
        } else {
            const headers = Array.from(node.querySelectorAll('tr:first-child th'))
                .map(th => clean(th.innerText));

            // Detect format: does tbody use <th scope="row"> for player name?
            const firstBodyRow = node.querySelector('tbody tr');
            const hasThPlayer = !!firstBodyRow?.querySelector('th[scope="row"]');

            const rows = Array.from(node.querySelectorAll('tbody tr')).map(tr => {
                let playerName = '';
                let tds;

                if (hasThPlayer) {
                    // Format B: player in <th scope="row">, rest in <td>
                    const th = tr.querySelector('th[scope="row"]');
                    playerName = th ? extractName(th) : '';
                    tds = Array.from(tr.querySelectorAll('td')).map(td => clean(td.innerText));
                } else {
                    // Format A: all <td>, player typically at index 2 (after Pos, No)
                    tds = Array.from(tr.querySelectorAll('td'));
                    // Find the td that has a .fn span or a data-sort-value (player cell)
                    const nameCell = tds.find(td => td.querySelector('.fn, [data-sort-value]'));
                    if (nameCell) {
                        playerName = extractName(nameCell);
                        tds = tds.map(td => clean(td.innerText));
                    } else if (tds.length > 2) {
                        playerName = extractName(tds[2]);
                        tds = tds.map(td => clean(td.innerText));
                    } else {
                        tds = tds.map(td => clean(td.innerText));
                    }
                }

                return { playerName, tds, hasThPlayer };
            }).filter(r => r.playerName);

            items.push({ type: 'table', headers, rows, hasThPlayer });
        }
    }
    return items;
}
"""


def scrape_wiki_year(page: Page, year: int) -> dict[str, list[dict]]:
    """Returns {nationality: [{player_name, team, league}, ...]}"""
    url = WIKI_URLS.get(year)
    if not url:
        return {}

    page.goto(url, wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(2000)

    items = page.evaluate(EXTRACT_JS)

    results: dict[str, list[dict]] = {}
    current_country = ""

    for item in items:
        if item["type"] == "heading":
            current_country = item["text"]
        elif item["type"] == "table" and current_country:
            nat = COUNTRY_MAP.get(current_country)
            if not nat:
                continue

            headers = item["headers"]
            rows = item["rows"]
            has_th = item.get("hasThPlayer", False)
            players = parse_roster_table(headers, rows, has_th)
            if players:
                if nat not in results:
                    results[nat] = []
                results[nat].extend(players)

    return results


def parse_roster_table(headers: list[str], rows: list[dict], has_th_player: bool) -> list[dict]:
    """
    Format B (has_th_player=True):  th=Player + tds=[No, Pos, DOB, Team, League, ...]
      headers = ['Player','No.','Pos.','DOB','Team','League',...]  → td_idx = header_idx - 1
    Format A (has_th_player=False): all tds=[Pos, No, Player, DOB, Bats, Throws, Club/Team]
      headers = ['Pos.','No.','Player','DOB','Bats','Throws','Club']  → td_idx = header_idx
    """
    players = []
    h_lower = [h.lower() for h in headers]

    if has_th_player:
        # Format B: player in th, td columns start at header index 1 → td index 0
        def td_idx(keyword: str) -> int | None:
            for i, h in enumerate(h_lower):
                if keyword in h and i >= 1:
                    return i - 1
            return None
    else:
        # Format A: all tds, direct header index mapping
        def td_idx(keyword: str) -> int | None:
            for i, h in enumerate(h_lower):
                if keyword in h:
                    return i
            return None

    has_club = any("club" in h for h in h_lower)
    team_td   = td_idx("club") if has_club else td_idx("team")
    league_td = td_idx("league")

    for row in rows:
        name = row.get("playerName", "").strip()
        tds  = row.get("tds", [])
        # Skip header rows and position-code-only names (e.g. "P", "C", "SS")
        if not name or name.lower() in ("player", "pos", "pos.") or re.fullmatch(r"[A-Z0-9]{1,2}", name):
            continue

        # Extract parenthetical CJK name before stripping: "Shohei Ohtani (大谷翔平)" → "大谷翔平"
        paren_match = re.search(r"\(([^\)]+)\)", name)
        name_zh = ""
        if paren_match:
            candidate = paren_match.group(1).strip()
            # Keep only if it contains CJK characters
            if re.search(r"[\u3000-\u9fff\uac00-\ud7af]", candidate):
                name_zh = candidate
        name = re.sub(r"\s*\(.*?\)\s*", "", name).strip()

        team   = tds[team_td].strip()   if team_td   is not None and len(tds) > team_td   else ""
        league = tds[league_td].strip() if league_td is not None and len(tds) > league_td else ""

        # Strip flag country name prepended to team (e.g. "United States San Diego Padres")
        team = re.sub(r"^[A-Z][a-zA-Z ]+(?:States?|Republic|Kingdom|Islands?|Rico|Korea|Taiwan|Japan|Cuba|Venezuela|Dominican|Panama|Nicaragua|Colombia|Mexico|Canada|Australia|Italy|Netherlands|Israel|Czechia|Britain|Zealand|Brazil|Africa|Germany|France|Spain|China|Chinese Taipei)\s+", "", team).strip()

        players.append({"player_name": name, "name_zh": name_zh, "team": team, "league": league})

    return players


def strip_accents(s: str) -> str:
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")


def title_hyphenated(s: str) -> str:
    """Title-case each hyphenated segment: 'kuan-yu' → 'Kuan-Yu'."""
    return "-".join(w.capitalize() for w in s.split("-"))


def flip_name(name: str) -> str:
    """'Chen Kuan-yu' → 'Kuan-Yu Chen'; 'C. C. Lee' → unchanged (3+ simple tokens)."""
    # Normalize spaces in initials: "C. C." → "C.C."
    name = re.sub(r"([A-Z])\.\s+([A-Z])\.", r"\1.\2.", name)
    parts = name.split()
    if len(parts) == 2:
        # "Family Given-name" → "Given-Name Family"
        return f"{title_hyphenated(parts[1])} {title_hyphenated(parts[0])}"
    return name


def name_variants(name: str) -> list[str]:
    """Return candidate normalized forms of a Wikipedia player name."""
    seen: list[str] = []

    def add(s: str) -> None:
        if s and s not in seen:
            seen.append(s)

    # Fix "C. C." spacing
    norm = re.sub(r"([A-Z])\.\s+([A-Z])\.", r"\1.\2.", name)
    add(norm)

    # Title-cased version (Kuan-yu → Kuan-Yu)
    tc = " ".join(title_hyphenated(p) for p in norm.split())
    add(tc)

    # Accent-stripped variants
    add(strip_accents(norm))
    add(strip_accents(tc))

    # Flipped (Family-First → Given-First): "Chen Kuan-yu" → "Kuan-Yu Chen"
    flipped = flip_name(norm)
    if flipped != norm:
        add(flipped)
        add(strip_accents(flipped))

    # Hyphen-to-space in given name: "Baek-Ho Kang" → "Baek Ho Kang"
    # Also try removing hyphen/space entirely: "Hye-seong Kim" → "Hyeseong Kim"
    for v in list(seen):
        no_hyph = v.replace("-", " ")
        if no_hyph != v:
            add(no_hyph)
            add(strip_accents(no_hyph))
        # collapse: remove hyphens and collapse internal spaces per token
        parts = v.split()
        collapsed = " ".join(p.replace("-", "") for p in parts)
        if collapsed != v:
            add(collapsed)
            add(strip_accents(collapsed))

    return seen


def update_sheet1(all_data: dict[int, dict[str, list[dict]]]) -> None:
    wb = load_workbook(OUTPUT)
    ws = wb["Sheet1"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    # Build index: (nationality, season, player_name) → row
    # Plain index: accent-stripped; Lower index: accent-stripped + lowercased
    row_index: dict[tuple, int] = {}
    row_index_plain: dict[tuple, int] = {}
    row_index_lower: dict[tuple, int] = {}
    for r in range(2, ws.max_row + 1):
        nat    = ws.cell(r, col["nationality"]).value
        season = ws.cell(r, col["season"]).value
        name   = ws.cell(r, col["player_name"]).value
        if nat and season and name:
            row_index[(nat, season, name)] = r
            plain = strip_accents(name)
            row_index_plain[(nat, season, plain)] = r
            row_index_lower[(nat, season, plain.lower())] = r

    updated = unmatched = 0

    for year, nat_data in all_data.items():
        for nat, players in nat_data.items():
            for p in players:
                r = None
                for variant in name_variants(p["player_name"]):
                    r = row_index.get((nat, year, variant))
                    if r is None:
                        plain_v = strip_accents(variant)
                        r = row_index_plain.get((nat, year, plain_v))
                    if r is None:
                        r = row_index_lower.get((nat, year, plain_v.lower()))
                    if r is not None:
                        break
                if r is None:
                    unmatched += 1
                    continue
                if p.get("team"):
                    ws.cell(r, col["team"]).value = p["team"]
                if p.get("league"):
                    ws.cell(r, col["league"]).value = p["league"]
                if p.get("name_zh") and not ws.cell(r, col["player_name_zh"]).value:
                    ws.cell(r, col["player_name_zh"]).value = p["name_zh"]
                updated += 1

    wb.save(OUTPUT)
    print(f"\nUpdated: {updated}  Unmatched: {unmatched}")


def main() -> None:
    all_data: dict[int, dict[str, list[dict]]] = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1440, "height": 900})

        for year in sorted(WIKI_URLS.keys(), reverse=True):
            print(f"\n[{year}] Scraping Wikipedia...")
            nat_data = scrape_wiki_year(page, year)
            all_data[year] = nat_data
            for nat, players in nat_data.items():
                print(f"  {nat}: {len(players)} players")

        browser.close()

    update_sheet1(all_data)
    print("Done.")


if __name__ == "__main__":
    main()
